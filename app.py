import streamlit as st
import sqlite3
import pandas as pd
import hashlib
import datetime
import time
import os
import cv2
import face_recognition
import numpy as np
from PIL import Image
import io
import base64
import random
import streamlit as st
import sqlite3
import hashlib
import pandas as pd
from datetime import datetime
from io import BytesIO
from PIL import Image
import base64

import socket
from pyqrcode import QRCode
import pyotp
from io import BytesIO
import base64
import streamlit.components.v1 as components
import time


import shutil
import os

if 'show_edit_form' not in st.session_state:
    st.session_state.show_edit_form = False
if 'editing_user_id' not in st.session_state:
    st.session_state.editing_user_id = None


from datetime import datetime, timedelta ,date
import pickle
import os

import mediapipe as mp
import math

mp_face_mesh = mp.solutions.face_mesh
face_mesh = mp_face_mesh.FaceMesh(
    max_num_faces=1,
    refine_landmarks=True,
    min_detection_confidence=0.5,
    min_tracking_confidence=0.5
)


SALES_USERNAME = "sales"
SALES_PASSWORD = "sales123" 


FACE_ENCODINGS_DIR = "face_encodings"


if not os.path.exists(FACE_ENCODINGS_DIR):
    os.makedirs(FACE_ENCODINGS_DIR)


if 'authenticated' not in st.session_state:
    st.session_state.authenticated = False
if 'user_id' not in st.session_state:
    st.session_state.user_id = None
if 'user_role' not in st.session_state:
    st.session_state.user_role = None
if 'username' not in st.session_state:
    st.session_state.username = None
if 'face_encodings_db' not in st.session_state:
    st.session_state.face_encodings_db = {}
if 'is_sales_login' not in st.session_state:
        st.session_state.is_sales_login = False
# Database setup
def init_db():
    conn = sqlite3.connect('attendance_system.db', check_same_thread=False)
    c = conn.cursor()
    
    c.execute('''
    CREATE TABLE IF NOT EXISTS announcements (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        title TEXT,
        message TEXT,
        created_at TEXT
    )
    ''')
    # Create users table
    c.execute('''
    CREATE TABLE IF NOT EXISTS users (
        id INTEGER PRIMARY KEY,
        username TEXT UNIQUE NOT NULL,
        password TEXT NOT NULL,
        role TEXT NOT NULL,
        name TEXT NOT NULL,
        email TEXT,
        phone TEXT,
        department TEXT NOT NULL,
        join_date TEXT,
        face_encoding BLOB
    )
    ''')

    # Create projects table
    c.execute('''
    CREATE TABLE IF NOT EXISTS projects (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL,
        description TEXT,
        link TEXT,
        members TEXT,
        deadline TEXT,
        total_hours REAL,
        status TEXT,
        client_name TEXT,
        client_phone TEXT,
        client_email TEXT,
        extra_details TEXT
    )
    ''')

    # Add photo column if not exists
    c.execute("PRAGMA table_info(users)")
    columns = [column[1] for column in c.fetchall()]
    if 'photo' not in columns:
        c.execute("ALTER TABLE users ADD COLUMN photo BLOB")
    
    # In the init_db() function, add this after the existing ALTER TABLE check for 'photo'
    c.execute("PRAGMA table_info(users)")
    columns = [column[1] for column in c.fetchall()]
    if 'professional_skills' not in columns:
        c.execute("ALTER TABLE users ADD COLUMN professional_skills TEXT")

    # Create attendance table
    c.execute('''
    CREATE TABLE IF NOT EXISTS attendance (
        id INTEGER PRIMARY KEY,
        user_id INTEGER NOT NULL,
        date TEXT NOT NULL,
        check_in TEXT,
        check_out TEXT,
        work_hours REAL,
        status TEXT,
        verification_method TEXT,
        FOREIGN KEY (user_id) REFERENCES users (id)
    )
    ''')

    # Create daily reports table
    c.execute('''
    CREATE TABLE IF NOT EXISTS daily_reports (
        id INTEGER PRIMARY KEY,
        user_id INTEGER NOT NULL,
        date TEXT NOT NULL,
        project TEXT,
        activities TEXT,
        achievements TEXT,
        challenges TEXT,
        FOREIGN KEY (user_id) REFERENCES users (id)
    )
    ''')

    # Create salary table
    c.execute('''
    CREATE TABLE IF NOT EXISTS salary (
        id INTEGER PRIMARY KEY,
        user_id INTEGER NOT NULL,
        month TEXT NOT NULL,
        year INTEGER NOT NULL,
        base_salary REAL NOT NULL,
        bonus REAL DEFAULT 0,
        deductions REAL DEFAULT 0,
        total_salary REAL NOT NULL,
        payment_status TEXT DEFAULT 'Pending',
        payment_date TEXT,
        FOREIGN KEY (user_id) REFERENCES users (id)
    )
    ''')
    # Create leave_requests table
    c.execute('''
    CREATE TABLE IF NOT EXISTS leave_requests (
        id INTEGER PRIMARY KEY,
        user_id INTEGER NOT NULL,
        start_date TEXT NOT NULL,
        end_date TEXT NOT NULL,
        leave_type TEXT NOT NULL,
        reason TEXT NOT NULL,
        status TEXT DEFAULT 'Pending',
        admin_approval TEXT,
        employee_verification TEXT,
        created_at TEXT DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (user_id) REFERENCES users (id)
    )
    ''')

    # Create an admin if no admin
    c.execute("SELECT COUNT(*) FROM users WHERE role='admin'")
    if c.fetchone()[0] == 0:
        admin_password = hashlib.sha256("admin123".encode()).hexdigest()
        c.execute(
            "INSERT INTO users (username, password, role, name, email, department, join_date) VALUES (?, ?, ?, ?, ?, ?, ?)",
            ("admin", admin_password, "admin", "Administrator", "admin@company.com", "Management",
             datetime.now().strftime("%Y-%m-%d")))

    conn.commit()
    conn.close()
# ------------------new college contact tab --------------
def college_details_tab():
    st.header("College Details Management")
    
    # Check and install required dependencies
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        st.warning("Installing required dependencies...")
        import subprocess
        import sys
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
        st.rerun()
    
    # Database setup for colleges
    conn = sqlite3.connect('attendance_system.db')
    c = conn.cursor()
    c.execute('''
        CREATE TABLE IF NOT EXISTS colleges (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            location TEXT NOT NULL,
            contact TEXT NOT NULL,
            details TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    conn.commit()
    conn.close()
    
    tab1, tab2 = st.tabs(["Add College", "Import Excel"])
    
    with tab1:
        with st.form("college_form", clear_on_submit=True):
            name = st.text_input("College Name*", placeholder="Enter college name")
            location = st.text_input("Location*", placeholder="City, State")
            contact = st.text_input("Contact Information*", placeholder="Phone/Email")
            details = st.text_area("Additional Details", placeholder="Any other information")
            
            submitted = st.form_submit_button("Add College")
            if submitted:
                if not name or not location or not contact:
                    st.warning("Please fill all required fields (*)")
                else:
                    conn = sqlite3.connect('attendance_system.db')
                    c = conn.cursor()
                    c.execute("""
                        INSERT INTO colleges (name, location, contact, details)
                        VALUES (?, ?, ?, ?)
                    """, (name, location, contact, details))
                    conn.commit()
                    conn.close()
                    st.success("College added successfully!")
    
    with tab2:
        st.subheader("Import Excel File")
        
        # Template download
        st.markdown("### Download Template")
        template_df = pd.DataFrame(columns=['name', 'location', 'contact', 'details'])
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            template_df.to_excel(writer, index=False, sheet_name='Colleges')
            workbook = writer.book
            worksheet = writer.sheets['Colleges']
            
            # Create style for header
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(
                start_color="4472C4",  # Corrected color format
                end_color="4472C4", 
                fill_type="solid"
            )
            
            # Apply style to header row
            for row in worksheet.iter_rows(min_row=1, max_row=1):
                for cell in row:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")
            
            # Set column widths
            worksheet.column_dimensions['A'].width = 30  # name
            worksheet.column_dimensions['B'].width = 25  # location
            worksheet.column_dimensions['C'].width = 25  # contact
            worksheet.column_dimensions['D'].width = 40  # details
        
        excel_data = output.getvalue()
        st.download_button(
            label="Download Excel Template",
            data=excel_data,
            file_name="college_template.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
        uploaded_file = st.file_uploader("Upload Excel file", type=["xlsx", "xls"])
        
        if uploaded_file:
            try:
                # Read the Excel file
                df = pd.read_excel(uploaded_file)
                
                # Rename columns to match database schema
                column_mapping = {
                    'College Name': 'name',
                    'College': 'name',
                    'Institution': 'name',
                    'Location': 'location',
                    'City': 'location',
                    'State': 'location',
                    'Contact Information': 'contact',
                    'Contact': 'contact',
                    'Phone': 'contact',
                    'Email': 'contact',
                    'Additional Info': 'details',
                    'Notes': 'details',
                    'Description': 'details'
                }
                
                df = df.rename(columns=column_mapping)
                
                # Validate required columns
                required_columns = ['name', 'location', 'contact']
                missing_columns = [col for col in required_columns if col not in df.columns]
                
                if missing_columns:
                    st.error(f"Missing required columns: {', '.join(missing_columns)}")
                else:
                    # Handle missing contact information
                    missing_contact = df['contact'].isna() | (df['contact'] == '')
                    
                    if missing_contact.any():
                        st.warning(f"Found {missing_contact.sum()} rows with missing contact information")
                        
                        # Option to fill missing values
                        fill_option = st.radio("How to handle missing contact information:",
                                              ["Set to 'Not provided'", "Skip these rows", "Cancel import"])
                        
                        if fill_option == "Set to 'Not provided'":
                            df.loc[missing_contact, 'contact'] = 'Not provided'
                        elif fill_option == "Skip these rows":
                            df = df[~missing_contact]
                        else:  # Cancel import
                            st.stop()
                    
                    # Add missing columns with default values
                    for col in required_columns + ['details']:
                        if col not in df.columns:
                            df[col] = ''
                    
                    # Connect to database
                    conn = sqlite3.connect('attendance_system.db')
                    
                    # Append to existing colleges
                    df[['name', 'location', 'contact', 'details']].to_sql(
                        'colleges', 
                        conn, 
                        if_exists='append', 
                        index=False
                    )
                    
                    # Get updated data
                    updated_df = pd.read_sql_query("SELECT * FROM colleges", conn)
                    conn.close()
                    
                    st.success(f"Successfully imported {len(df)} colleges!")
                    
                    # Show updated data
                    st.subheader("All Colleges")
                    st.dataframe(
                        updated_df.drop(columns=['id']), 
                        use_container_width=True,
                        hide_index=True
                    )
                    
                    # Download updated Excel file
                    output = BytesIO()
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        updated_df.to_excel(writer, index=False, sheet_name='Colleges')
                        workbook = writer.book
                        worksheet = writer.sheets['Colleges']
                        
                        # Create style for header
                        header_font = Font(bold=True, color="FFFFFF")
                        header_fill = PatternFill(
                            start_color="4472C4", 
                            end_color="4472C4", 
                            fill_type="solid"
                        )
                        
                        # Apply style to header row
                        for row in worksheet.iter_rows(min_row=1, max_row=1):
                            for cell in row:
                                cell.font = header_font
                                cell.fill = header_fill
                                cell.alignment = Alignment(horizontal="center")
                        
                        # Auto-adjust column widths
                        for col_idx, column in enumerate(worksheet.columns, 1):
                            max_length = 0
                            column_letter = get_column_letter(col_idx)
                            for cell in column:
                                try:
                                    if len(str(cell.value)) > max_length:
                                        max_length = len(str(cell.value))
                                except:
                                    pass
                            adjusted_width = (max_length + 2)
                            worksheet.column_dimensions[column_letter].width = adjusted_width
                    
                    # Prepare download
                    excel_data = output.getvalue()
                    st.download_button(
                        label="Download Updated Excel File",
                        data=excel_data,
                        file_name="updated_college_details.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
            except Exception as e:
                st.error(f"Error processing file: {str(e)}")
# ------------------end new college contact tab ------------
# Add custom CSS and JS
def inject_custom_files():
    with open("assets/style.css") as f:
        st.markdown(f"<style>{f.read()}</style>", unsafe_allow_html=True)
        
    with open("assets/script.js") as f:
        st.markdown(f"<script>{f.read()}</script>", unsafe_allow_html=True)


inject_custom_files()

# User authentication
def authenticate(username, password):
    conn = sqlite3.connect('attendance_system.db')
    c = conn.cursor()

    hashed_password = hashlib.sha256(password.encode()).hexdigest()
    c.execute("SELECT id, role, name FROM users WHERE username = ? AND password = ?", (username, hashed_password))
    result = c.fetchone()

    conn.close()

    if result:
        return {'id': result[0], 'role': result[1], 'name': result[2]}
    return None


def load_face_encodings():
    """Load all face encodings from pickle files"""
    face_encodings = {}
    try:
        for file_name in os.listdir(FACE_ENCODINGS_DIR):
            if file_name.endswith(".pkl"):
                user_id = int(file_name.split("_")[1].split(".")[0])
                with open(os.path.join(FACE_ENCODINGS_DIR, file_name), 'rb') as f:
                    face_encodings[user_id] = pickle.load(f)
    except Exception as e:
        st.error(f"Error loading face encodings: {e}")
    return face_encodings


def save_face_encoding(user_id, face_encoding):
    """Save face encoding to a pickle file"""
    try:
        
        FACE_ENCODINGS_DIR = "face_encodings"

        
        if not os.path.exists(FACE_ENCODINGS_DIR):
            os.makedirs(FACE_ENCODINGS_DIR)

        
        file_path = os.path.join(FACE_ENCODINGS_DIR, f"user_{user_id}.pkl")

        
        with open(file_path, 'wb') as f:
            pickle.dump(face_encoding, f)

        
        return True
    except Exception as e:
        
        st.error(f"Error saving face encoding: {e}")
        return False

def register_face(user_id):
    """Capture multiple angles and store all encodings"""
    encodings = []
    directions = ['front', 'left', 'right']
    
    for direction in directions:
        st.write(f"Face {direction} and click capture")
        img_file = st.camera_input(f"Face {direction}", key=f"cam_{direction}")
        
        if img_file:
            image = Image.open(img_file)
            image_array = np.array(image)
            
            
            face_encodings = face_recognition.face_encodings(image_array)
            if face_encodings:
                encodings.extend(face_encodings)  
                
    if len(encodings) < 3:
        st.error("Failed to capture sufficient angles")
        return False
    
    
    with open(os.path.join(FACE_ENCODINGS_DIR, f"user_{user_id}.pkl"), 'wb') as f:
        pickle.dump(encodings, f)
    
    return True

def verify_face(image_array, user_id):
    """Verify against all stored encodings"""
    try:
        with open(os.path.join(FACE_ENCODINGS_DIR, f"user_{user_id}.pkl"), 'rb') as f:
            stored_encodings = pickle.load(f)
    except:
        return False
    
    current_encoding = face_recognition.face_encodings(image_array)
    if not current_encoding:
        return False
    
    
    matches = face_recognition.compare_faces(stored_encodings, current_encoding[0], tolerance=0.4)
    return any(matches)

def mark_attendance(user_id, verification_method):
    conn = sqlite3.connect('attendance_system.db')
    c = conn.cursor()

    today = datetime.now().strftime("%Y-%m-%d")
    current_time = datetime.now().strftime("%H:%M:%S")

    
    c.execute("SELECT check_in, check_out FROM attendance WHERE user_id = ? AND date = ?", (user_id, today))
    result = c.fetchone()

    if result is None:
        
        c.execute(
            "INSERT INTO attendance (user_id, date, check_in, status, verification_method) VALUES (?, ?, ?, ?, ?)",
            (user_id, today, current_time, "Present", verification_method))
        message = f"Check-in recorded at {current_time}"
    elif result[1] is None:
        
        check_in_time = datetime.strptime(result[0], "%H:%M:%S")
        check_out_time = datetime.strptime(current_time, "%H:%M:%S")

        
        work_duration = check_out_time - check_in_time
        work_hours = round(work_duration.total_seconds() / 3600, 2) 

        c.execute("UPDATE attendance SET check_out = ?, work_hours = ? WHERE user_id = ? AND date = ?",
                  (current_time, work_hours, user_id, today))
        message = f"Check-out recorded at {current_time}. You worked for {work_hours} hours today."
    else:
        
        message = "You have already completed your attendance for today."

    conn.commit()
    conn.close()
    return message


def get_user_attendance(user_id, start_date=None, end_date=None):
    conn = sqlite3.connect('attendance_system.db')

    if start_date and end_date:
        query = f"""
        SELECT a.date, a.check_in, a.check_out, a.work_hours, a.status, a.verification_method
        FROM attendance a
        WHERE a.user_id = ? AND a.date BETWEEN ? AND ?
        ORDER BY a.date DESC
        """
        df = pd.read_sql_query(query, conn, params=(user_id, start_date, end_date))
    else:
        query = f"""
        SELECT a.date, a.check_in, a.check_out, a.work_hours, a.status, a.verification_method
        FROM attendance a
        WHERE a.user_id = ?
        ORDER BY a.date DESC
        """
        df = pd.read_sql_query(query, conn, params=(user_id,))

    conn.close()
    return df


def get_all_attendance(start_date=None, end_date=None):
    conn = sqlite3.connect('attendance_system.db')

    if start_date and end_date:
        query = f"""
        SELECT u.name, a.date, a.check_in, a.check_out, a.work_hours, a.status, a.verification_method
        FROM attendance a
        JOIN users u ON a.user_id = u.id
        WHERE a.date BETWEEN ? AND ?
        ORDER BY a.date DESC, u.name
        """
        df = pd.read_sql_query(query, conn, params=(start_date, end_date))
    else:
        query = f"""
        SELECT u.name, a.date, a.check_in, a.check_out, a.work_hours, a.status, a.verification_method
        FROM attendance a
        JOIN users u ON a.user_id = u.id
        ORDER BY a.date DESC, u.name
        """
        df = pd.read_sql_query(query, conn)

    conn.close()
    return df



def submit_daily_report(user_id, project, activities, achievements, challenges):
    conn = sqlite3.connect('attendance_system.db')
    c = conn.cursor()
    today = datetime.now().strftime("%Y-%m-%d")

    
    c.execute("SELECT id FROM daily_reports WHERE user_id = ? AND date = ? AND project = ?", (user_id, today, project))
    result = c.fetchone()

    if result:
        
        c.execute("""UPDATE daily_reports 
                     SET activities = ?, achievements = ?, challenges = ? 
                     WHERE id = ?""",
                  (activities, achievements, challenges, result[0]))
        message = "Your daily report has been updated."
    else:
        
        c.execute("""INSERT INTO daily_reports 
                     (user_id, date, project, activities, achievements, challenges) 
                     VALUES (?, ?, ?, ?, ?, ?)""",
                  (user_id, today, project, activities, achievements, challenges))
        message = "Your daily report has been submitted."

    conn.commit()
    conn.close()
    return message


def get_user_reports(user_id, start_date=None, end_date=None, projects=None):
    conn = sqlite3.connect("attendance_system.db")

    query = """
        SELECT date, project, activities, achievements, challenges
        FROM daily_reports
        WHERE user_id = ?
    """
    params = [user_id]

    if start_date and end_date:
        query += " AND date BETWEEN ? AND ?"
        params += [start_date, end_date]

    if projects:
        placeholders = ','.join('?' for _ in projects)
        query += f" AND project IN ({placeholders})"
        params += projects

    query += " ORDER BY date DESC"

    df = pd.read_sql_query(query, conn, params=params)
    conn.close()
    return df



def get_all_reports(start_date=None, end_date=None, projects=None):
    conn = sqlite3.connect('attendance_system.db')

    query = """
        SELECT u.name, r.date, r.project, r.activities, r.achievements, r.challenges
        FROM daily_reports r
        JOIN users u ON r.user_id = u.id
        WHERE 1=1
    """
    params = []

    if start_date and end_date:
        query += " AND r.date BETWEEN ? AND ?"
        params += [start_date, end_date]

    if projects:
        placeholders = ','.join('?' for _ in projects)
        query += f" AND r.project IN ({placeholders})"
        params += projects

    query += " ORDER BY r.date DESC, u.name"

    df = pd.read_sql_query(query, conn, params=params)
    conn.close()
    return df


def add_salary(user_id, month, year, base_salary, bonus=0, deductions=0):
    conn = sqlite3.connect('attendance_system.db')
    c = conn.cursor()

    total_salary = base_salary + bonus - deductions

    
    c.execute("SELECT id FROM salary WHERE user_id = ? AND month = ? AND year = ?", (user_id, month, year))
    result = c.fetchone()

    if result:
        
        c.execute("""
        UPDATE salary SET base_salary = ?, bonus = ?, deductions = ?, total_salary = ? 
        WHERE id = ?
        """, (base_salary, bonus, deductions, total_salary, result[0]))
        message = f"Salary record updated for {month} {year}"
    else:
        
        c.execute("""
        INSERT INTO salary (user_id, month, year, base_salary, bonus, deductions, total_salary)
        VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (user_id, month, year, base_salary, bonus, deductions, total_salary))
        message = f"Salary record created for {month} {year}"

    conn.commit()
    conn.close()
    return message


def process_payment(salary_id):
    conn = sqlite3.connect('attendance_system.db')
    c = conn.cursor()

    payment_date = datetime.now().strftime("%Y-%m-%d")
    c.execute("UPDATE salary SET payment_status = 'Paid', payment_date = ? WHERE id = ?", (payment_date, salary_id))

    conn.commit()
    conn.close()
    return "Payment processed successfully"


def get_user_salary(user_id):
    conn = sqlite3.connect('attendance_system.db')

    query = f"""
    SELECT id, month, year, base_salary, bonus, deductions, total_salary, payment_status, payment_date
    FROM salary
    WHERE user_id = ?
    ORDER BY year DESC, CASE 
        WHEN month = 'January' THEN 1
        WHEN month = 'February' THEN 2
        WHEN month = 'March' THEN 3
        WHEN month = 'April' THEN 4
        WHEN month = 'May' THEN 5
        WHEN month = 'June' THEN 6
        WHEN month = 'July' THEN 7
        WHEN month = 'August' THEN 8
        WHEN month = 'September' THEN 9
        WHEN month = 'October' THEN 10
        WHEN month = 'November' THEN 11
        WHEN month = 'December' THEN 12
    END DESC
    """
    df = pd.read_sql_query(query, conn, params=(user_id,))

    conn.close()
    return df


def get_all_salary(month=None, year=None):
    conn = sqlite3.connect('attendance_system.db')

    if month and year:
        query = f"""
        SELECT s.id, u.name, s.month, s.year, s.base_salary, s.bonus, s.deductions, s.total_salary, s.payment_status, s.payment_date
        FROM salary s
        JOIN users u ON s.user_id = u.id
        WHERE s.month = ? AND s.year = ?
        ORDER BY u.name
        """
        df = pd.read_sql_query(query, conn, params=(month, year))
    else:
        query = f"""
        SELECT s.id, u.name, s.month, s.year, s.base_salary, s.bonus, s.deductions, s.total_salary, s.payment_status, s.payment_date
        FROM salary s
        JOIN users u ON s.user_id = u.id
        ORDER BY s.year DESC, CASE 
            WHEN s.month = 'January' THEN 1
            WHEN s.month = 'February' THEN 2
            WHEN s.month = 'March' THEN 3
            WHEN s.month = 'April' THEN 4
            WHEN s.month = 'May' THEN 5
            WHEN s.month = 'June' THEN 6
            WHEN s.month = 'July' THEN 7
            WHEN s.month = 'August' THEN 8
            WHEN s.month = 'September' THEN 9
            WHEN s.month = 'October' THEN 10
            WHEN s.month = 'November' THEN 11
            WHEN s.month = 'December' THEN 12
        END DESC, u.name
        """
        df = pd.read_sql_query(query, conn)

    conn.close()
    return df



def get_all_users():
    try:
        conn = sqlite3.connect('attendance_system.db')
        query = """
        SELECT id, username, name, email, phone, department, join_date, role, photo
        FROM users
        ORDER BY name
        """
        users_df = pd.read_sql_query(query, conn)
        conn.close()
        return users_df
    except Exception as e:
        st.error(f"Error fetching users: {str(e)}")
        return pd.DataFrame()


def delete_user(user_id):
    try:
        conn = sqlite3.connect('attendance_system.db')
        c = conn.cursor()
        c.execute("DELETE FROM users WHERE id = ?", (user_id,))
        conn.commit()
        conn.close()
        return True
    except Exception as e:
        st.error(f"Error deleting user: {str(e)}")
        return False

def update_user(user_id, name, email, phone, department, role, join_date=None):
    try:
        user_id = int(user_id) 
        conn = sqlite3.connect('attendance_system.db')
        c = conn.cursor()
        
        c.execute("""
        UPDATE users SET name = ?, email = ?, phone = ?, department = ?, role = ?, join_date = ?
        WHERE id = ?
        """, (name, email, phone, department, role, join_date, user_id))
        
        conn.commit()
        conn.close()
        return True, "User updated successfully"
    except sqlite3.Error as e:
        return False, f"Database error: {str(e)}"
    except Exception as e:
        return False, f"Error updating user: {str(e)}"


def change_password(user_id, new_password):
    try:
        conn = sqlite3.connect('attendance_system.db')
        c = conn.cursor()
        
        hashed_password = hashlib.sha256(new_password.encode()).hexdigest()
        c.execute("UPDATE users SET password = ? WHERE id = ?", (hashed_password, user_id))
        
        conn.commit()
        conn.close()
        return True, "Password changed successfully"
    except sqlite3.Error as e:
        return False, f"Password change failed: {str(e)}"


def get_user_info(user_id):
    conn = sqlite3.connect('attendance_system.db')
    c = conn.cursor()
    
    c.execute("""
    SELECT username, name, email, phone, department, join_date, role, photo
    FROM users WHERE id = ?
    """, (user_id,))
    
    result = c.fetchone()
    conn.close()
    
    if result:
        return {
            'username': result[0],
            'name': result[1],
            'email': result[2],
            'phone': result[3],
            'department': result[4],
            'join_date': result[5],
            'role': result[6],
            'photo': result[7]
        }
    return None

def get_user_role(user_id):
    try:
        conn = sqlite3.connect('attendance_system.db')
        c = conn.cursor()
        c.execute("SELECT role FROM users WHERE id = ?", (user_id,))
        role = c.fetchone()[0]
        conn.close()
        return role
    except:
        return None

def display_user_image(photo_bytes):
    if photo_bytes:
        try:
            image = Image.open(BytesIO(photo_bytes))
            return image
        except:
            return None
    return None

st.markdown("""
<style>
    .user-card {
        border: 1px solid #ddd;
        border-radius: 5px;
        padding: 10px;
        margin-bottom: 10px;
        background-color: #f9f9f9;
    }
    .required-field::after {
        content: "*";
        color: red;
        margin-left: 2px;
    }
    .user-image {
        border-radius: 50%;
        object-fit: cover;
    }
    .st-emotion-cache-16idsys p {
        font-weight: bold;     
    }
    .delete-btn {
        color: white;
        background-color: #ff4b4b;
        border-radius: 5px;
        padding: 2px 8px;
        border: none;
        cursor: pointer;
    }
    .admin-badge {
        background-color: #4CAF50;
        color: white;
        padding: 2px 8px;
        border-radius: 5px;
        font-size: 0.8em;
    }
    .employee-badge {
        background-color: #2196F3;
        color: white;
        padding: 2px 8px;
        border-radius: 5px;
        font-size: 0.8em;
    }
</style>
""", unsafe_allow_html=True)

def generate_attendance_report(month, year):
    conn = sqlite3.connect('attendance_system.db')

    
    if month == 1:
        first_day = f"{year}-01-01"
        last_day = f"{year}-01-31"
    elif month == 2:
        if year % 4 == 0 and (year % 100 != 0 or year % 400 == 0): 
            last_day = 29
        else:
            last_day = 28
        first_day = f"{year}-02-01"
        last_day = f"{year}-02-{last_day}"
    elif month in [4, 6, 9, 11]:
        first_day = f"{year}-{month:02d}-01"
        last_day = f"{year}-{month:02d}-30"
    else:
        first_day = f"{year}-{month:02d}-01"
        last_day = f"{year}-{month:02d}-31"

    query = f"""
    SELECT u.name, u.department, 
           COUNT(CASE WHEN a.status = 'Present' THEN 1 END) as present_days,
           COUNT(DISTINCT a.date) as total_days,
           SUM(a.work_hours) as total_hours
    FROM users u
    LEFT JOIN attendance a ON u.id = a.user_id AND a.date BETWEEN ? AND ?
    WHERE u.role = 'employee'
    GROUP BY u.id
    ORDER BY u.name
    """

    df = pd.read_sql_query(query, conn, params=(first_day, last_day))

    conn.close()
    return df


# UI Functions

def login_page():
    st.markdown(
    "<h1 style='text-align: center; color: white; margin-left: 2rem;'> WorkLense </h1>", 
    unsafe_allow_html=True
)


    # Inject custom CSS and JS
    st.markdown("""
    <style>
        @import url('https://fonts.googleapis.com/css2?family=Poppins:wght@300;400;600&display=swap');
        
        * {
            font-family: 'Poppins', sans-serif;
        }
                
         body, h1, h2, h3, p, div {
            color: #a0a0a0;
        }
        
        /* Specific element targeting */
        .stMarkdown h1,  /* "Employee Attendance System" */
        .stMarkdown h2,  /* "Secure Login" */
        .stMarkdown h3,  /* "Login Method" */
        .stMarkdown p,   /* Paragraph texts */
        .stAlert,        /* Deprecation warning */
        .stCameraInput > label /* Camera instructions */ {
            color: #a0a0a0;
        }
        
        .main > div {
            background-color: transparent !important;
            padding-top: 0 !important;
        }
        
        body::before {
            content: '';
            position: fixed;
            top: 0;
            left: 0;
            right: 0;
            bottom: 0;
            z-index: -1;
            background-image: url('https://images.unsplash.com/photo-1497366754035-f200968a6e72?ixlib=rb-1.2.1&auto=format&fit=crop&w=1350&q=80');
            background-size: cover;
            background-position: center;
            filter: blur(8px) brightness(0.8);
        }
        
        .login-container {
            background: rgba(255, 255, 255, 0.85);
            border: 1px solid rgba(0, 0, 0, 0.2);
            border-radius: 15px;
            box-shadow: 0 8px 32px 0 rgba(31, 38, 135, 0.15);
            padding: 2rem;
            margin: 2rem auto;
            max-width: 500px;
            backdrop-filter: blur(5px);
        }
        
        .login-title {
            color: white;
            text-align: center;
            margin-bottom: 1.5rem;
            font-size: 1.8rem;
            font-weight: 600;
        }
        
        .stLoginContainer {
            background: rgba(255, 255, 255, 0.95) !important;
            backdrop-filter: blur(12px);
            border-radius: 20px;
            box-shadow: 0 8px 32px 0 rgba(31, 38, 135, 0.37);
            padding: 2rem;
            animation: slideUp 0.5s ease;
            border: 1px solid rgba(255, 255, 255, 0.18);
            margin: auto;
            max-width: 500px;
        }
        
        .stRadio > div {
            flex-direction: row !important;
            gap: 0.3rem;
        }
        
        .stRadio label {
            background: rgba(240, 242, 246, 0.3) !important;
            border-radius: 10px !important;
            padding: 0.5rem 1rem !important;
            transition: all 0.3s ease !important;
            border: 1px solid rgba(255, 255, 255, 0.2) !important;
        }
        
                
        .stRadio label:hover {
            transform: translateY(-2px);
            box-shadow: 0 4px 15px rgba(0,0,0,0.1);
        }
                
        .stTextInput input, .stTextInput input:focus {
            border-radius: 10px !important;
            padding: 0.8rem !important;
            border: 2px solid rgba(224, 224, 224, 0.3) !important;
            background: rgba(255, 255, 255, 0.1) !important;
            color: black !important;
        }
        
        .stButton button {
            width: 100% !important;
            border-radius: 10px !important;
            background: linear-gradient(45deg, #000, #000) !important;
            color: white !important;
            font-weight: bold !important;
            transition: all 0.3s ease !important;
            border: none !important;
            padding: 0.8rem !important;
        }
        
        .logo-container {
            display: flex;
            justify-content: center;
            align-items: center;
            height: 100%;
            padding: 2rem;
        }
        
        .logo-image {
            max-width: 300px;
            filter: drop-shadow(0 4px 8px rgba(0,0,0,0.1));
        }
        
        @keyframes slideUp {
            from { transform: translateY(50px); opacity: 0; }
            to { transform: translateY(0); opacity: 1; }
        }
    </style>
    
    <script>
    // Tilt effect for the system overview image
    document.addEventListener('DOMContentLoaded', (event) => {
        const image = document.querySelector('.system-image img');
        if(image) {
            image.addEventListener('mousemove', (e) => {
                const rect = image.getBoundingClientRect();
                const x = (e.clientX - rect.left) / image.width;
                const y = (e.clientY - rect.top) / image.height;
                
                image.style.transform = `perspective(1000px) 
                    rotateX(${(0.5 - y) * 10}deg) 
                    rotateY(${(x - 0.5) * 10}deg) 
                    scale(1.05)`;
            });
            
            image.addEventListener('mouseleave', () => {
                image.style.transform = 'perspective(1000px) rotateX(0) rotateY(0) scale(1)';
            });
        }
    });
    </script>
    """, unsafe_allow_html=True)

    col1, col2 = st.columns([2, 1])

    with col1:
        st.markdown("<h3 style='text-align: center; color: white; margin-left: 2rem;'>🔐 Secure Login</h3>", 
                    unsafe_allow_html=True)

        login_method = st.radio("Login Method", ["Password", "Biometric", "QR Code"], horizontal=True)

        if login_method == "Password":
            username = st.text_input("Username")
            password = st.text_input("Password", type="password")

            if st.button("Login"):
                if username and password:
                    if username == SALES_USERNAME and password == SALES_PASSWORD:
                        st.session_state.authenticated = True
                        st.session_state.is_sales_login = True
                        st.session_state.username = SALES_USERNAME
                        st.success(f"Welcome, Sales User!")
                        st.rerun()
                    else:
                        user = authenticate(username, password)
                        if user:
                            st.session_state.authenticated = True
                            st.session_state.user_id = user['id']
                            st.session_state.user_role = user['role']
                            st.session_state.username = username
                            st.session_state.is_sales_login = False
                            st.success(f"Welcome, {user['name']}!")
                            st.rerun()
                        else:
                            st.error("Invalid username or password")
                else:
                    st.warning("Please enter both username and password")

        elif login_method == "Biometric":
            
            st.write("Look at the camera for face verification")
            img_file = st.camera_input("Take a photo for face verification", key="login_camera")

            if img_file is not None:
                image = Image.open(img_file)
                image_array = np.array(image)

                
                face_encodings_db = load_face_encodings()

                
                face_locations = face_recognition.face_locations(image_array)
                if len(face_locations) != 1:
                    st.error("Please ensure only one face is in the frame")
                    return

                
                current_encoding = face_recognition.face_encodings(image_array, face_locations)[0]

                
                matches = []
                for user_id, saved_encoding in face_encodings_db.items():
                    match = face_recognition.compare_faces([saved_encoding], current_encoding, tolerance=0.5)
                    if match[0]:
                        matches.append(user_id)

                if matches:
                    
                    user_id = matches[0]

                    
                    conn = sqlite3.connect('attendance_system.db')
                    c = conn.cursor()
                    c.execute("SELECT id, username, role, name FROM users WHERE id = ?", (user_id,))
                    user = c.fetchone()
                    conn.close()

                    if user:
                        st.session_state.authenticated = True
                        st.session_state.user_id = user[0]
                        st.session_state.username = user[1]
                        st.session_state.user_role = user[2]
                        st.success(f"Welcome, {user[3]}!")

                        
                        if user[2] == 'admin':
                            st.session_state.is_admin = True
                        else:
                            st.session_state.is_admin = False
                        
                        st.rerun()
                    else:
                        st.error("User not found in the database")
                else:
                    st.error("No matching face found. Please try again or use password login.")

        elif login_method == "QR Code":
            st.subheader("QR Code Login")
            qr_img, secret = generate_qr_code()
            
            st.image(BytesIO(base64.b64decode(qr_img)), 
                   caption="Scan this QR code with your mobile device")
            
            if 'qr_countdown' not in st.session_state:
                st.session_state.qr_countdown = 300

            st.session_state.qr_countdown -= 1
            mins, secs = divmod(st.session_state.qr_countdown, 60)
            st.write(f"QR code refreshes in: {mins:02d}:{secs:02d}")
            
            token = st.query_params.get("token", "")
            if token:
                totp = pyotp.TOTP(secret)
                if totp.verify(token, valid_window=1):
                    st.success("QR Code verified!")
                    st.session_state.authenticated = True
                    st.session_state.user_role = "employee"
                    st.rerun()
                else:
                    st.error("Invalid or expired QR code")
            
            if st.session_state.qr_countdown <= 0:
                st.session_state.qr_countdown = 300
                st.rerun()
            
            time.sleep(1)
            st.rerun()

    with col2:
        st.markdown('<div class="system-image">', unsafe_allow_html=True)
        st.image(
            r"D:\Sudhi\sudhi project\entudio\Logo1-removebg-preview.png",
            width=400,
           use_container_width='auto',
        )
        st.markdown('</div>', unsafe_allow_html=True)


def get_all_salary_info(month, year):
    """Get all salary information for a specific month and year"""
    conn = sqlite3.connect('attendance_system.db')
    query = """
    SELECT u.name, s.month, s.year, s.base_salary, s.bonus, s.deductions, 
           s.total_salary, s.payment_status, s.payment_date
    FROM salary s
    JOIN users u ON s.user_id = u.id
    WHERE s.month = ? AND s.year = ?
    ORDER BY u.name
    """
    salary_df = pd.read_sql_query(query, conn, params=(month, year))
    conn.close()

    return salary_df


def get_monthly_attendance_summary(user_id, month, year):
    """Get attendance summary for a specific user for a month"""
    conn = sqlite3.connect('attendance_system.db')
    c = conn.cursor()

    
    import calendar
    last_day = calendar.monthrange(year, month)[1]
    start_date = f"{year}-{month:02d}-01"
    end_date = f"{year}-{month:02d}-{last_day}"

    
    c.execute("""
    SELECT COUNT(*), SUM(work_hours)
    FROM attendance
    WHERE user_id = ? AND date BETWEEN ? AND ? AND status = 'Present'
    """, (user_id, start_date, end_date))

    result = c.fetchone()
    present_days = result[0] if result[0] else 0
    work_hours = result[1] if result[1] else 0

    conn.close()

    return {
        'present_days': present_days,
        'work_hours': work_hours
    }


def save_salary_info(user_id, month, year, base_salary, bonus, deductions, total_salary, payment_status):
    """Save or update salary information for a user"""
    conn = sqlite3.connect('attendance_system.db')
    c = conn.cursor()
    
    
    c.execute("""
    SELECT id FROM salary
    WHERE user_id = ? AND month = ? AND year = ?
    """, (user_id, month, year))
    
    existing = c.fetchone()
    payment_date = datetime.now().strftime("%Y-%m-%d") if payment_status == "Paid" else None
    
    if existing:
        
        c.execute("""
        UPDATE salary
        SET base_salary = ?, bonus = ?, deductions = ?, total_salary = ?,
            payment_status = ?, payment_date = ?
        WHERE user_id = ? AND month = ? AND year = ?
        """, (base_salary, bonus, deductions, total_salary, payment_status, 
              payment_date, user_id, month, year))
    else:
        
        c.execute("""
        INSERT INTO salary (user_id, month, year, base_salary, bonus, deductions,
                            total_salary, payment_status, payment_date)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (user_id, month, year, base_salary, bonus, deductions, 
              total_salary, payment_status, payment_date))
    
    conn.commit()
    conn.close()
    
    return True

def get_salary_history(month=None, year=None):
    """Get salary history with optional month/year filter"""
    conn = sqlite3.connect('attendance_system.db')
    
    query = """
    SELECT u.name as employee_name, s.month, s.year, s.base_salary, s.bonus,
           s.deductions, s.total_salary, s.payment_status, s.payment_date
    FROM salary s
    JOIN users u ON s.user_id = u.id
    """
    
    params = []
    where_clauses = []
    
    
    if month:
        month_num = {
            "January": 1, "February": 2, "March": 3, "April": 4,
            "May": 5, "June": 6, "July": 7, "August": 8,
            "September": 9, "October": 10, "November": 11, "December": 12
        }.get(month)
        if month_num:
            where_clauses.append("s.month = ?")
            params.append(month_num)
    
    if year and year != "All":
        where_clauses.append("s.year = ?")
        params.append(year)
    
    if where_clauses:
        query += " WHERE " + " AND ".join(where_clauses)
    
    query += " ORDER BY s.year DESC, s.month DESC, u.name"
    
    salary_df = pd.read_sql_query(query, conn, params=params)
    conn.close()
    
    
    month_names = {
        1: "January", 2: "February", 3: "March", 4: "April",
        5: "May", 6: "June", 7: "July", 8: "August",
        9: "September", 10: "October", 11: "November", 12: "December"
    }
    
    if not salary_df.empty:
        salary_df['month_name'] = salary_df['month'].apply(
            lambda x: month_names.get(int(x), "") if pd.notnull(x) and str(x).isdigit() else ""
        )
    
    return salary_df

def logout():
    for key in st.session_state.keys():
        del st.session_state[key]
    st.rerun()


def employee_dashboard():
    st.title(f"Employee Dashboard - {st.session_state.username}")
    
    show_announcement_popup() 
    
    user_info = get_user_info(st.session_state.user_id)
    
    

    menu = st.sidebar.selectbox(
        "Menu",
        ["Home", "Mark Attendance", "View Attendance", "Daily Report", "Salary Information", "Leave Management", "Profile",
         "Biometric Setup","Allocated Projects", "Annoncement"]
    )

    if st.sidebar.button("Logout"):
        logout()

    if menu == "Home":
        col1, col2 = st.columns([2, 1])

        with col1:
            st.header("Welcome to Employee Portal")
            st.markdown(f"""
            ### {user_info['name']}
            - **Department:** {user_info['department']}
            - **Joined:** {user_info['join_date']}
            """)

            
            today = datetime.now().strftime("%Y-%m-%d")
            conn = sqlite3.connect('attendance_system.db')
            c = conn.cursor()
            c.execute("SELECT check_in, check_out FROM attendance WHERE user_id = ? AND date = ?",
                      (st.session_state.user_id, today))
            today_attendance = c.fetchone()
            conn.close()

            st.subheader("Today's Status")
            if today_attendance:
                if today_attendance[1]: 
                    st.success(f"✅ Checked in at {today_attendance[0]} and checked out at {today_attendance[1]}")
                else:
                    st.info(f"⏳ Checked in at {today_attendance[0]}, not checked out yet")
            else:
                st.warning("❌ Not checked in today")

        with col2:
            st.subheader("Quick Actions")

            
            if today_attendance and today_attendance[1]:
                st.button("Already Completed Today", disabled=True)
            elif today_attendance:
                if st.button("Check Out"):
                    message = mark_attendance(st.session_state.user_id, "Manual")
                    st.success(message)
                    st.rerun()
            else:
                if st.button("Check In"):
                    message = mark_attendance(st.session_state.user_id, "Manual")
                    st.success(message)
                    st.rerun()

        st.subheader("Attendance Summary (Last 7 Days)")
        end_date = datetime.now()
        start_date = end_date - timedelta(days=7)
        recent_attendance = get_user_attendance(
            st.session_state.user_id,
            start_date.strftime("%Y-%m-%d"),
            end_date.strftime("%Y-%m-%d")
        )
        if not recent_attendance.empty:
            st.dataframe(recent_attendance, use_container_width=True)
        else:
            st.info("No recent attendance records found")

        
        st.subheader("Latest Daily Report")
        recent_reports = get_user_reports(
            st.session_state.user_id,
            start_date.strftime("%Y-%m-%d"),
            end_date.strftime("%Y-%m-%d")
        )
        if not recent_reports.empty:
            latest_report = recent_reports.iloc[0]
            st.markdown(f"**Date:** {latest_report['date']}")
            st.markdown(f"**Activities:** {latest_report['activities']}")
            st.markdown(f"**Achievements:** {latest_report['achievements']}")
            st.markdown(f"**Challenges:** {latest_report['challenges']}")
        else:
            st.info("No recent daily reports found")

    elif menu == "Mark Attendance":
        st.header("Mark Attendance")

        today = datetime.now().strftime("%Y-%m-%d")
        current_time = datetime.now().strftime("%H:%M:%S")

        st.markdown(f"**Date:** {today} | **Time:** {current_time}")

        
        conn = sqlite3.connect('attendance_system.db')
        c = conn.cursor()
        c.execute("SELECT check_in, check_out FROM attendance WHERE user_id = ? AND date = ?",
                  (st.session_state.user_id, today))
        today_attendance = c.fetchone()
        conn.close()

        col1, col2 = st.columns(2)

        with col1:
            st.subheader("Attendance Method")
            attendance_method = st.radio("Select verification method:",
                                         ["Manual", "Biometric (Face Recognition)"],
                                         horizontal=True)

            if attendance_method == "Biometric (Face Recognition)":
                
                if st.session_state.face_encodings_db and st.session_state.user_id in st.session_state.face_encodings_db:

                    img_file = st.camera_input("Look at the camera for face verification", key="attendance_camera")

                    
                    if img_file is not None:
                        st.write("Processing face...")

                        try:
                            
                            image = Image.open(img_file)
                            image_array = np.array(image)

                            
                            is_verified = verify_face(image_array, st.session_state.user_id)
                            if is_verified:
                                message = mark_attendance(st.session_state.user_id, "Biometric")
                                st.success(f"Face verified! {message}")
                                st.rerun()
                            else:
                                st.error("Face verification failed. Please try again or use manual check-in.")
                        except Exception as e:
                            st.error(f"Error processing image: {str(e)}")
                    
                else:
                    st.warning("You need to set up your face data in the Biometric Setup section first.")

            
            if attendance_method == "Manual":
                if today_attendance and today_attendance[1]:
                    st.info("You have already completed attendance for today.")
                elif today_attendance:
                    if st.button("Check Out"):
                        message = mark_attendance(st.session_state.user_id, "Manual")
                        st.success(message)
                        st.rerun()
                else:
                    if st.button("Check In"):
                        message = mark_attendance(st.session_state.user_id, "Manual")
                        st.success(message)
                        st.rerun()

        with col2:
            st.subheader("Attendance Status")
            if today_attendance:
                if today_attendance[1]:  # Checked out
                    st.success(f"✅ Checked in at {today_attendance[0]} and checked out at {today_attendance[1]}")
                else:
                    st.info(f"⏳ Checked in at {today_attendance[0]}, not checked out yet")
            else:
                st.warning("❌ Not checked in today")

    elif menu == "View Attendance":
        st.header("My Attendance Records")

        
        col1, col2 = st.columns(2)
        with col1:
                start_date = st.date_input("From", value=datetime.now() - timedelta(days=7))
        with col2:
                end_date = st.date_input("To", value=datetime.now())

        if start_date > end_date:
            st.error("End date must be after start date")
        else:
            
            attendance_data = get_user_attendance(
                st.session_state.user_id,
                start_date.strftime("%Y-%m-%d"),
                end_date.strftime("%Y-%m-%d")
            )

            if not attendance_data.empty:
                
                total_days = (end_date - start_date).days + 1
                present_days = attendance_data.shape[0]
                total_work_hours = attendance_data['work_hours'].sum()

                
                col1, col2, col3 = st.columns(3)
                with col1:
                    st.metric("Total Working Days", total_days)
                with col2:
                    st.metric("Days Present", present_days)
                with col3:
                    st.metric("Total Work Hours", f"{total_work_hours:.2f}")

                
                st.subheader("Attendance Details")
                st.dataframe(attendance_data, use_container_width=True)

                
                csv = attendance_data.to_csv(index=False)
                st.download_button(
                    label="Download as CSV",
                    data=csv,
                    file_name=f"attendance_{start_date}_to_{end_date}.csv",
                    mime="text/csv"
                )
            else:
                st.info("No attendance records found for the selected date range")
    elif menu == "Annoncement":
        st.subheader("📢 Company Announcements")

        conn = sqlite3.connect('attendance_system.db')
        c = conn.cursor()
        c.execute("SELECT title, message, created_at FROM announcements ORDER BY id DESC")
        announcements = c.fetchall()
        conn.close()

        if announcements:
            for title, message, created_at in announcements:
                with st.container():
                    st.markdown(
                        f"""
                        <div style="
                            background-color: #f0f2f6;
                            padding: 15px 20px;
                            border-radius: 10px;
                            box-shadow: 0 2px 5px rgba(0,0,0,0.1);
                            margin-bottom: 15px;
                        ">
                            <h4 style="margin-bottom: 5px; color: #2c3e50;"> {title}</h4>
                            <p style="margin: 0; font-size: 14px; color: #555;">{message}</p>
                            <p style="text-align: right; font-size: 12px; color: #888;">🕒 {created_at}</p>
                        </div>
                        """,
                        unsafe_allow_html=True
                    )
        else:
            st.info("No announcements available.")
    elif menu == "Daily Report":
        st.header("Daily Reports")
        tab1, tab2 = st.tabs(["Submit Report", "View Reports"])

        
        with tab1:
            st.subheader("Submit Daily Report")
            conn = sqlite3.connect("attendance_system.db")
            c = conn.cursor()
            c.execute("SELECT name FROM projects")
            project_names = [row[0] for row in c.fetchall()]
            conn.close()

            project_options = ["General"] + project_names
            selected_projects = st.multiselect("Select at least one Project", options=project_options)

            if not selected_projects:
                st.warning("⚠️ Please select at least one project.")

            today = datetime.now().strftime("%Y-%m-%d")
            st.markdown(f"**Date:** {today}")
            activities = st.text_area("Activities Performed Today", height=150)
            achievements = st.text_area("Achievements", height=100)
            challenges = st.text_area("Challenges Faced", height=100)

            if st.button("Submit Report"):
                if not selected_projects:
                    st.error("❌ Report not submitted. Select at least one project.")
                else:
                    project_str = ", ".join(selected_projects)
                    conn = sqlite3.connect("attendance_system.db")
                    c = conn.cursor()
                    user_id = st.session_state.user_id

                    
                    c.execute("SELECT id FROM daily_reports WHERE user_id = ? AND date = ?", (user_id, today))
                    existing = c.fetchone()

                    if existing:
                        c.execute("""
                            UPDATE daily_reports
                            SET project = ?, activities = ?, achievements = ?, challenges = ?
                            WHERE id = ?
                        """, (project_str, activities, achievements, challenges, existing[0]))
                    else:
                        c.execute("""
                            INSERT INTO daily_reports
                            (user_id, date, project, activities, achievements, challenges)
                            VALUES (?, ?, ?, ?, ?, ?)
                        """, (user_id, today, project_str, activities, achievements, challenges))

                    conn.commit()
                    conn.close()
                    st.success(f"✅ Report submitted for {today} with projects: {project_str}")

        
        with tab2:
            st.header("My Daily Reports")
            conn = sqlite3.connect("attendance_system.db")
            c = conn.cursor()
            c.execute("SELECT DISTINCT project FROM daily_reports WHERE user_id = ?", (st.session_state.user_id,))
            projects = list({p.strip() for row in c.fetchall() for p in row[0].split(",")})
            conn.close()

            selected_projects = st.multiselect("Select Project(s)", options=projects)
            col1, col2 = st.columns(2)
            with col1:
                start_date = st.date_input("From", value=datetime.now() - timedelta(days=30))
            with col2:
                end_date = st.date_input("To", value=datetime.now())

            if start_date > end_date:
                st.error("End date must be after start date.")
            else:
                conn = sqlite3.connect("attendance_system.db")
                query = "SELECT date, project, activities, achievements, challenges FROM daily_reports WHERE user_id = ? AND date BETWEEN ? AND ?"
                params = [st.session_state.user_id, start_date.strftime("%Y-%m-%d"), end_date.strftime("%Y-%m-%d")]

                if selected_projects:
                    query += " AND (" + " OR ".join(["project LIKE ?"] * len(selected_projects)) + ")"
                    params.extend([f"%{proj}%" for proj in selected_projects])

                df = pd.read_sql_query(query, conn, params=params)
                conn.close()

                if not df.empty:
                    for _, row in df.iterrows():
                        with st.expander(f"Report: {row['date']} ({row['project']})"):
                            st.markdown(f"**Activities:** {row['activities']}")
                            st.markdown(f"**Achievements:** {row['achievements'] or 'None'}")
                            st.markdown(f"**Challenges:** {row['challenges'] or 'None'}")

                    csv = df.to_csv(index=False)
                    st.download_button("Download All Reports as CSV", data=csv, file_name="my_reports.csv", mime="text/csv")
                else:
                    st.info("No reports found for selected filters.")



    elif menu == "Leave Management":
        st.header("Leave Management")

        tab1, tab2, tab3 = st.tabs(["New Leave Request", "My Leave History", "Total Leave Days"])

        with tab1:
            st.subheader("Submit Leave Request")
            with st.form("leave_form"):
                col1, col2 = st.columns(2)
                with col1:
                    start_date = st.date_input("Start Date")
                    leave_type = st.selectbox("Leave Type", ["Sick", "Vacation", "Personal", "Other"])
                with col2:
                    end_date = st.date_input("End Date")
                    reason = st.text_area("Reason")

                if st.form_submit_button("Submit Request"):
                    if start_date > end_date:
                        st.error("End date must be after start date")
                    else:
                        create_leave_request(
                            st.session_state.user_id,
                            start_date.strftime("%Y-%m-%d"),
                            end_date.strftime("%Y-%m-%d"),
                            leave_type,
                            reason
                        )
                        st.success("Leave request submitted for admin approval")

        with tab2:
            st.subheader("Leave History")
            leaves = get_employee_leave_requests(st.session_state.user_id)
            if not leaves.empty:
                for _, leave in leaves.iterrows():
                    with st.expander(f"{leave['start_date']} to {leave['end_date']} - {leave['status']}"):
                        col1, col2 = st.columns(2)
                        with col1:
                            st.markdown(f"**Type:** {leave['leave_type']}")
                            st.markdown(f"**Reason:** {leave['reason']}")
                        with col2:
                            st.markdown(f"**Status:** {leave['status']}")
                            if leave['admin_approval']: 
                                st.markdown(f"**Admin Note:** {leave['admin_approval']}")
                                if leave['status'] == 'Approved' and not leave['employee_verification']:
                                    if st.button("Verify Leave", key=f"verify_{leave['id']}"):
                                        verify_leave_request(leave['id'])
                                        st.success("Leave verified!")
            else:
                st.info("No leave requests found")

            with tab3:
                st.subheader("Total Leave Days")
                total_leaves = get_all_leave_requests_by_user(st.session_state.user_id)
                if not total_leaves.empty and "status" in total_leaves.columns:
                    approved_leaves = total_leaves[total_leaves["status"] == "Approved"]
                    if not approved_leaves.empty:
                        approved_leaves["leave_days"] = (
                            pd.to_datetime(approved_leaves["end_date"]) - pd.to_datetime(approved_leaves["start_date"])
                        ).dt.days + 1
                        total_days = approved_leaves["leave_days"].sum()
                        st.metric("Total Approved Leave Days", total_days)
                        st.dataframe(approved_leaves[["start_date", "end_date", "leave_type", "leave_days"]])
                    else:
                        st.info("No approved leave records found.")
                else:
                    st.info("No leave records found.")


    elif menu == "Salary Information":
        st.header("My Salary Information")

        salary_data = get_user_salary(st.session_state.user_id)

        if not salary_data.empty:
            st.subheader("Salary History")

            
            for _, row in salary_data.iterrows():
                with st.expander(f"{row['month']} {row['year']} - {row['payment_status']}"):
                    col1, col2 = st.columns(2)
                    with col1:
                        st.markdown(f"**Base Salary:** {row['base_salary']:.2f}")
                        st.markdown(f"**Bonus:** {row['bonus']:.2f}")
                        st.markdown(f"**Deductions:** {row['deductions']:.2f}")
                    with col2:
                        st.markdown(f"**Total Salary:** {row['total_salary']:.2f}")
                        st.markdown(f"**Payment Status:** {row['payment_status']}")
                        if row['payment_date']:
                            st.markdown(f"**Payment Date:** {row['payment_date']}")
        else:
            st.info("No salary records found")

   
    elif menu == "Profile":
        st.header("My Profile")

        if user_info:
            col1, col2, col3 = st.columns(3)  

            with col1:
                st.subheader("Personal Information")
                st.markdown(f"**Name:** {user_info['name']}")
                st.markdown(f"**Username:** {user_info['username']}")
                st.markdown(f"**Email:** {user_info['email']}")
                st.markdown(f"**Phone:** {user_info['phone']}")
                st.markdown(f"**Department:** {user_info['department']}")
                st.markdown(f"**Join Date:** {user_info['join_date']}")

            with col2:
                st.subheader("Change Password")
                current_password = st.text_input("Current Password", type="password")
                new_password = st.text_input("New Password", type="password")
                confirm_password = st.text_input("Confirm New Password", type="password")

                if st.button("Update Password"):
                    if not current_password or not new_password or not confirm_password:
                        st.warning("Please fill all password fields")
                    elif new_password != confirm_password:
                        st.error("New passwords do not match")
                    else:
                        
                        hashed_current = hashlib.sha256(current_password.encode()).hexdigest()
                        conn = sqlite3.connect('attendance_system.db')
                        c = conn.cursor()
                        c.execute("SELECT id FROM users WHERE id = ? AND password = ?",
                                (st.session_state.user_id, hashed_current))
                        user_verified = c.fetchone()
                        conn.close()

                        if user_verified:
                            message = change_password(st.session_state.user_id, new_password)
                            st.success(message)
                        else:
                            st.error("Current password is incorrect")

            with col3:
                st.subheader("Professional Skills")
                
                
                conn = sqlite3.connect('attendance_system.db')
                c = conn.cursor()
                c.execute("SELECT professional_skills FROM users WHERE id = ?", (st.session_state.user_id,))
                current_skills = c.fetchone()[0] or ""
                conn.close()
                
                
                new_skills = st.text_area(
                    "Your Skills (comma separated)",
                    value=current_skills,
                    placeholder="e.g., Python, SQL, Project Management"
                )
                
                if st.button("Update Skills"):
                    try:
                        conn = sqlite3.connect('attendance_system.db')
                        c = conn.cursor()
                        c.execute(
                            "UPDATE users SET professional_skills = ? WHERE id = ?",
                            (new_skills, st.session_state.user_id)
                        )
                        conn.commit()
                        conn.close()
                        st.success("Professional skills updated successfully!")
                        st.rerun()
                    except Exception as e:
                        st.error(f"Error updating skills: {str(e)}")

    elif menu == "Biometric Setup":
        st.header("Biometric Setup")

        st.markdown("""
        ### Face Recognition Setup
        Use this section to register or update your facial data for biometric login and attendance.
        """)

        
        if st.session_state.face_encodings_db and st.session_state.user_id in st.session_state.face_encodings_db:
            st.info("You already have facial data registered. You can update it below.")
        else:
            st.warning("No facial data registered yet. Please use the camera below to register your face.")

        st.write("For best results, ensure good lighting and position your face clearly in the frame.")
        face_image = st.camera_input("Take a photo for face registration", key="face_reg_camera")

        if face_image is not None:
            image = Image.open(face_image)
            image_array = np.array(image)
            face_locations = face_recognition.face_locations(image_array)

            if not face_locations:
                st.error("No face detected in the image. Please try again with better lighting and positioning.")
            else:
                if len(face_locations) > 1:
                    st.warning("Multiple faces detected. Using the most prominent face.")
                
                face_encoding = face_recognition.face_encodings(image_array, [face_locations[0]])[0]
                
                
                if save_face_encoding(st.session_state.user_id, face_encoding):
                    st.session_state.face_encodings_db = load_face_encodings()
                    st.success("Face registered successfully! You can now use face verification.")

    elif menu == "Allocated Projects":

        st.header("My Allocated Projects")

        # Get current employee username

        user_info = get_user_info(st.session_state.user_id)
        employee_name = user_info['name'].strip().lower()

        # Get all projects

        projects = get_projects()

        # Filter projects where this employee is assigned

        my_projects = []

        for p in projects:

            members = [m.strip().lower() for m in p[4].split(",") if m.strip()]  # p[4] = members

            if employee_name in members:
                my_projects.append(p)

        if not my_projects:

            st.info("❌ You have no projects assigned.")

        else:

            for p in my_projects:
                with st.expander(f"{p[1]}"):
                    st.markdown(f"**Client Name:** {p[8]}")

                    st.markdown(f"**Client Phone:** {p[9]}")

                    st.markdown(f"**Client Email:** {p[10]}")

                    st.markdown(f"**Extra Details:** {p[11]}")

                    st.markdown(f"**Members:** {p[4]}")

                    st.markdown(f"**Deadline:** {p[5]}")

                    st.markdown(f"**Total Hours:** {p[6]}")

                    st.markdown(f"**Status:** {p[7]}")

                    st.markdown(f"[🔗 Project Link]({p[3]})")

def show_leave_notifications():
    if 'dismissed_notifications' not in st.session_state:
        st.session_state.dismissed_notifications = set()

    if 'notification_timer' not in st.session_state:
        st.session_state.notification_timer = {}

    
    conn = sqlite3.connect('attendance_system.db')
    c = conn.cursor()
    c.execute("""
        SELECT lr.id, u.name, lr.start_date, lr.end_date, lr.leave_type, lr.reason
        FROM leave_requests lr
        JOIN users u ON lr.user_id = u.id
        WHERE lr.status = 'Pending'
        ORDER BY lr.created_at DESC
        LIMIT 1
    """)
    req = c.fetchone()
    conn.close()

    if req:
        leave_id, name, start, end, leave_type, reason = req

        if leave_id in st.session_state.dismissed_notifications:
            return

        if leave_id not in st.session_state.notification_timer:
            st.session_state.notification_timer[leave_id] = time.time()

        
        if time.time() - st.session_state.notification_timer[leave_id] > 60:
            st.session_state.dismissed_notifications.add(leave_id)
            return

        
        html_content = f"""
        <div id="popup" style="
            position: fixed;
            bottom: 30px;
            right: 30px;
            background-color: #25D366;
            color: white;
            padding: 15px;
            border-radius: 10px;
            box-shadow: 0 4px 10px rgba(0,0,0,0.2);
            z-index: 9999;
            width: 300px;
            font-family: Arial, sans-serif;
            animation: fadeInUp 0.5s ease;
        ">
            <div style="display: flex; justify-content: space-between;">
                <strong>📝 New Leave Request</strong>
                <button onclick="document.getElementById('popup').remove(); fetch('/?dismiss_id={leave_id}')" 
                        style="background: transparent; color: white; border: none; font-size: 20px; cursor: pointer;">×</button>
            </div>
            <div style="margin-top: 10px;">
                <b>{name}</b> requested <b>{leave_type}</b> leave<br>
                📅 {start} → {end}<br>
                📌 Reason: {reason}
            </div>
        </div>

        <style>
        @keyframes fadeInUp {{
            from {{ opacity: 0; transform: translateY(20px); }}
            to {{ opacity: 1; transform: translateY(0); }}
        }}
        </style>
        """

        
        components.html(html_content, height=200)

        
        dismiss_id = st.query_params.get("dismiss_id")
        if dismiss_id:
            try:
                st.session_state.dismissed_notifications.add(int(dismiss_id))
                st.experimental_set_query_params()
                st.rerun()
            except:
                pass

def admin_dashboard():
    st.title("Admin Dashboard")
    
    show_leave_notifications()

    if st.session_state.get("user_role") != "admin":
        st.error("Only admins can access this section.")
        st.stop()
    #
    user_info = get_user_info(st.session_state.user_id)

    menu = st.sidebar.selectbox(
        "Menu",
        ["Home","User Management", "Attendance Management", "Leave Management", "Reports", "Project Reports", "Salary Management", "Employee Directory","College Details","College Details","Admin Directory","Admin Biometric Setup", 
         "Settings","QR Attendance"],index=["Home", "User Management", "Attendance Management","Leave Management","Reports","Project Reports", "Salary Management",  "Employee Directory","College Details","Admin Directory","Admin Biometric Setup", "Settings","QR Attendance"].index(
            st.session_state.get("menu", "Home")
        )
    )
    st.session_state.menu = menu 

    if st.sidebar.button("Logout"):
        logout()

    if menu == "Home":
        st.header(f"Welcome, {user_info['name']}")

        col1, col2 = st.columns(2)

        with col1:
            st.subheader("System Overview")

            conn = sqlite3.connect('attendance_system.db')
            c = conn.cursor()

            c.execute("SELECT COUNT(*) FROM users WHERE role = 'employee'")
            employee_count = c.fetchone()[0]

            today = datetime.now().strftime("%Y-%m-%d")
            c.execute("SELECT COUNT(*) FROM attendance WHERE date = ?", (today,))
            today_attendance = c.fetchone()[0]

            c.execute("SELECT COUNT(*) FROM daily_reports WHERE date = ?", (today,))
            today_reports = c.fetchone()[0]

            conn.close()

            st.metric("Total Employees", employee_count)
            st.metric("Today's Attendance", today_attendance)
            st.metric("Today's Reports", today_reports)

        with col2:
            st.subheader("Quick Actions")
            btn_col1, btn_col2 = st.columns(2)

            button_style = """
                <style>
                div.stButton > button {
                    width: 100%;
                }
                </style>
            """
            st.markdown(button_style, unsafe_allow_html=True)

            with btn_col1:
                if st.button("View Today's Attendance", key="btn_att"):
                    st.session_state.menu = "Attendance Management"
                    st.rerun()
                if st.button("Generate Monthly Report", key="btn_report"):
                    st.session_state.menu = "Reports"
                    st.rerun()

            with btn_col2:
                if st.button("Add New User", key="btn_add"):
                    st.session_state.menu = "User Management"
                    st.session_state.user_management_tab = "Add User"
                    st.rerun()
                if st.button("Open Sales Dashboard", key="btn_sales"):
                    st.session_state.is_sales_login = True
                    st.rerun()
    elif menu == "Admin Biometric Setup":
        st.header(" Admin Face Registration")
        
        if st.session_state.get("user_role") != "admin":
            st.error("Only admins can access this section.")
        else:
            st.markdown("""
            ### Register Your Face for Biometric Login
            Please position your face in front of the camera.
            """)

            img_file = st.camera_input("Capture your face")

            if img_file is not None:
                image = Image.open(img_file)
                image_array = np.array(image)

                face_locations = face_recognition.face_locations(image_array)

                if not face_locations:
                    st.error("No face detected. Try again.")
                elif len(face_locations) > 1:
                    st.warning("Multiple faces detected. Please ensure only one face is in the frame.")
                else:
                    face_encoding = face_recognition.face_encodings(image_array, face_locations)[0]

                    if save_face_encoding(st.session_state.user_id, face_encoding):
                        st.success("✅ Face registered successfully! You can now use biometric login.")


    elif menu == "QR Attendance":
                    st.header("QR Code Attendance")
 
                    qr_img, secret = generate_qr_code()

                    st.image(BytesIO(base64.b64decode(qr_img)), 
                            caption="Scan this QR code with your mobile device")

                    if 'qr_countdown' not in st.session_state:
                        st.session_state.qr_countdown = 300 

                    st.session_state.qr_countdown -= 1
                    mins, secs = divmod(st.session_state.qr_countdown, 60)
                    st.write(f"QR code refreshes in: {mins:02d}:{secs:02d}")

                    if st.session_state.qr_countdown <= 0:
                        st.session_state.qr_countdown = 300
                        st.rerun()
                    
                    time.sleep(1)
                    st.rerun()


    elif menu == "Admin Directory":
        display_user_directory("Admin Directory", "admin")
    
    # In the admin_dashboard() function, add this after the other menu options:
    elif menu == "College Details":
         college_details_tab()

    elif menu == "Employee Directory":
        display_user_directory("Employee Directory", "employee")

    elif menu == "User Management":
        st.header("User Management")

        tab_labels = ["All Users", "Add User", "Edit User"]
        default_tab = st.session_state.get("user_management_tab", "All Users")

        selected_tab = st.radio("Select a tab", tab_labels, index=tab_labels.index(default_tab), horizontal=True)

        if selected_tab == "All Users":
            st.subheader("All Users")
            users_df = get_all_users()
            
            if not users_df.empty:
                
                col1, col2 = st.columns([2, 1])
                with col1:
                    search_query = st.text_input("Search users", placeholder="Enter name, email, or department")
                with col2:
                    filter_dept = st.selectbox("Filter by department", 
                                            ["All"] + list(users_df['department'].unique()), 
                                            index=0)
                
                
                filtered_df = users_df.copy()
                if search_query:
                    filtered_df = filtered_df[
                        filtered_df['name'].str.contains(search_query, case=False) |
                        filtered_df['email'].str.contains(search_query, case=False) |
                        filtered_df['department'].str.contains(search_query, case=False)
                    ]
                
                if filter_dept != "All":
                    filtered_df = filtered_df[filtered_df['department'] == filter_dept]
                
                if not filtered_df.empty:
                    st.write(f"Showing {len(filtered_df)} users")
                    
                    
                    st.dataframe(
                        filtered_df[['id', 'username', 'name', 'email', 'phone', 'department', 'join_date', 'role']],
                        column_config={
                            "id": st.column_config.NumberColumn("ID", width="small"),
                            "username": st.column_config.TextColumn("Username"),
                            "name": st.column_config.TextColumn("Full Name"),
                            "email": st.column_config.TextColumn("Email"),
                            "phone": st.column_config.TextColumn("Phone", width="medium"),
                            "department": st.column_config.TextColumn("Department"),
                            "join_date": st.column_config.DateColumn("Join Date"),
                            "role": st.column_config.TextColumn("Role", width="small"),
                        },
                        use_container_width=True,
                        hide_index=True
                    )
                    
                    st.subheader("User Details & Actions")
                    for i, row in filtered_df.iterrows():
                        with st.container():
                            col1, col2, col3, col4 = st.columns([1, 2, 2, 1])
                            
                            with col1:
                                
                                if row['photo']:
                                    user_img = display_user_image(row['photo'])
                                    if user_img:
                                        st.image(user_img, width=80, caption=f"ID: {row['id']}")
                                    else:
                                        st.image("https://via.placeholder.com/80", width=80, caption=f"ID: {row['id']}")
                                else:
                                    st.image("https://via.placeholder.com/80", width=80, caption=f"ID: {row['id']}")
                            
                            with col2:
                                st.markdown(f"**{row['name']}** ({row['username']})")
                                st.markdown(f"📧 {row['email'] if row['email'] else 'No email'}")
                                st.markdown(f"📱 {row['phone'] if row['phone'] else 'No phone'}")
                            
                            with col3:
                                st.markdown(f"**Department:** {row['department']}")
                                st.markdown(f"**Joined:** {row['join_date']}")
                                role_badge = "admin-badge" if row['role'] == "admin" else "employee-badge"
                                st.markdown(f"<span class='{role_badge}'>{row['role'].upper()}</span>", unsafe_allow_html=True)
                            
                            with col4:
                                
                                action_col1, action_col2 = st.columns(2)
                                
                                
                                #if action_col1.button("✏️ Edit", key=f"edit_{row['id']}"):
                                    #st.session_state.editing_user_id = row['id']
                                    #st.session_state.show_edit_form = True
                                
                                # Delete button (not available for admins)
                                if row['role'] != 'admin':
                                        if st.button("Delete 🗑️", key=f"delete_{row['id']}"):
                                          if delete_user(row['id']):
                                               st.success(f"User {row['username']} deleted successfully!")
                                               st.rerun()    
                                else:
                                    action_col2.markdown(
                                             "<div><span style='color:gray'>Protected</span></div>",
                                                unsafe_allow_html=True)
                        
                        st.divider()
                    
                    
                    if 'show_edit_form' in st.session_state and st.session_state.show_edit_form:
                        user_id = st.session_state.editing_user_id
                        user_info = get_user_info(user_id)
                        
                        if user_info:
                            st.subheader(f"Edit User: {user_info['name']}")
                            
                            with st.form(key=f"edit_user_form_{user_id}"):
                                col1, col2 = st.columns(2)
                                
                                with col1:
                                    username = st.text_input("Username", user_info['username'], disabled=True)
                                    name = st.text_input("Full Name*", user_info['name'])
                                    email = st.text_input("Email", user_info['email'] if user_info['email'] else "")
                                
                                with col2:
                                    phone = st.text_input("Phone*", user_info['phone'] if user_info['phone'] else "", max_chars=10)
                                    department = st.selectbox("Department*", ["Software", "Hardware", "Sales", "Management", "Operation"], 
                                                        index=["Software", "Hardware", "Sales", "Management", "Operation"].index(user_info['department']) 
                                                        if user_info['department'] in ["Software", "Hardware", "Sales", "Management", "Operation"] else 0)
                                    role = st.selectbox("Role*", ["admin", "employee"], 
                                                    index=["admin", "employee"].index(user_info['role']) 
                                                    if user_info['role'] in ["admin", "employee"] else 1)
                                    
                                
                                
                                col1, col2 = st.columns(2)
                                update_button = col1.form_submit_button("Update User")
                                cancel_button = col2.form_submit_button("Cancel")
                                
                                if cancel_button:
                                    st.session_state.show_edit_form = False
                                    st.rerun()
                                
                                if update_button:
                                    errors = []
                                    
                                    
                                    if not name: errors.append("Full Name is required")
                                    if not phone: errors.append("Phone is required")
                                    if phone and (not phone.isdigit() or len(phone) != 10): 
                                        errors.append("Phone must be 10 digits")
                                    
                                    if errors:
                                        for error in errors:
                                            st.error(error)
                                    else:
                                        
                                        success, message = update_user(
                                            user_id, name, email, phone, department, role, user_info['join_date']
                                        )
                                        
                                        
                    
                    export_df = filtered_df.drop(columns=['photo'], errors='ignore')
                    csv = export_df.to_csv(index=False)
                    st.download_button(
                        label="📥 Download User List",
                        data=csv,
                        file_name="users_list.csv",
                        mime="text/csv",
                        key="user_list_download"
                    )
                else:
                    st.info("No users match your search criteria")
            else:
                st.info("No users found in the system")

        
        elif selected_tab == "Add User":
            st.subheader("Add New User")
            
            
            with st.form("add_user_form", clear_on_submit=True):
                
                col1, col2 = st.columns(2)
                
                with col1:
                    
                    st.markdown("<p class='required-field'>Username</p>", unsafe_allow_html=True)
                    username = st.text_input("", placeholder="Enter username", key="username_input")
                    
                    st.markdown("<p class='required-field'>Password</p>", unsafe_allow_html=True)
                    password = st.text_input("", placeholder="Enter password", type="password", key="password_input")
                    
                    st.markdown("<p class='required-field'>Full Name</p>", unsafe_allow_html=True)
                    name = st.text_input("", placeholder="Enter full name", key="name_input")
                    
                    email = st.text_input("Email", placeholder="user@company.com", key="email_input")
                
                with col2:
                    st.markdown("<p class='required-field'>Phone</p>", unsafe_allow_html=True)
                    phone = st.text_input("", placeholder="Enter 10-digit phone number", max_chars=10, key="phone_input")
                    
                    st.markdown("<p class='required-field'>Department</p>", unsafe_allow_html=True)
                    department = st.selectbox("", ["Software", "Hardware", "Sales", "Management", "Operation"], key="department_input")
                    
                    st.markdown("<p class='required-field'>Role</p>", unsafe_allow_html=True)
                    role = st.selectbox("", ["employee", "admin"], key="role_input")
                    
                    st.markdown("<p class='required-field'>Joining Date</p>", unsafe_allow_html=True)
                    join_date = st.date_input("", value=datetime.now(), key="date_input")
                
                
                st.markdown("<p>Profile Photo</p>", unsafe_allow_html=True)
                photo = st.file_uploader("", type=["jpg", "jpeg", "png"], key="photo_input")
                
                st.markdown("**Fields marked with * are required**")
                
                submitted = st.form_submit_button("Add User", use_container_width=True)
                
                if submitted:
                    errors = []
                    if not username: errors.append("Username is required")
                    if not password: errors.append("Password is required")
                    if not name: errors.append("Full Name is required")
                    if not phone: errors.append("Phone number is required")
                    if phone and (not phone.isdigit() or len(phone) != 10): 
                        errors.append("Phone must be 10 digits")
                    
                    if errors:
                        for error in errors:
                            st.error(error)
                    else:
                        
                        photo_bytes = None
                        if photo is not None:
                            photo_bytes = photo.getvalue()
                        
                        
                        success, user_id, message = add_user(
                            username, password, role, name, 
                            email, phone, department, join_date, photo_bytes
                        )
                        
                        if success:
                            st.success(message)
                            st.rerun()
                        else:
                            st.error(message)

            
            with st.expander("Need help? See an example of a properly filled form"):
                st.markdown("""
                **Example User Information:**
                - **Username**: john_doe
                - **Password**: Secure password (at least 8 characters)
                - **Full Name**: John Doe
                - **Email**: john.doe@company.com
                - **Phone**: 1234567890 (10 digits)
                - **Department**: Software
                - **Role**: employee (or admin for administrative access)
                - **Photo**: Upload a square image for best results
                """)
        elif selected_tab == "Edit User":
            st.subheader("Edit User")
            import os
            db_path = os.path.abspath('attendance_system.db')

            users_df = get_all_users()
            if not users_df.empty:
                user_to_edit = st.selectbox("Select User", users_df['username'].tolist(), key="user_to_edit_selectbox")

                selected_user = users_df[users_df['username'] == user_to_edit]
                if not selected_user.empty:
                    selected_user = selected_user.iloc[0]
                    
                    user_id = int(selected_user['id'])
                    st.subheader(f"Edit User (User ID: {user_id})") 

                    col1, col2 = st.columns(2)

                    with col1:
                        name = st.text_input("Full Name", value=selected_user['name'])
                        email = st.text_input("Email", value=selected_user['email'])
                        phone = st.text_input("Phone", value=selected_user['phone'])

                    with col2:
                        department = st.text_input("Department", value=selected_user['department'])

                        try:
                            default_date = datetime.strptime(selected_user['join_date'], "%Y-%m-%d").date() if \
                                selected_user['join_date'] else datetime.now().date()
                        except:
                            default_date = datetime.now().date()

                        join_date = st.date_input("Joining Date", value=default_date, key=f"join_date_{user_id}")

                        role = st.selectbox("Role", ["employee", "admin"], key="role_selectbox",
                                            index=0 if selected_user['role'] == "employee" else 1)
                        reset_password = st.checkbox("Reset Password")

                        if reset_password:
                            new_password = st.text_input("New Password", type="password")

                    if st.button("Update User", key="update_user_button"):
                        if not name.strip():
                            st.error("Name field cannot be empty")
                        else:
                            
                            def update_user_with_path(user_id, name, email, phone, department, role, join_date):
                                try:
                                    conn = sqlite3.connect(db_path)
                                    c = conn.cursor()

                                    c.execute("SELECT id, username FROM users WHERE id = ?", (user_id,))
                                    before_update = c.fetchone()

                                    c.execute("""
                                    UPDATE users SET name = ?, email = ?, phone = ?, department = ?, join_date = ?, role = ?
                                    WHERE id = ?
                                    """, (name, email, phone, department, join_date, role, user_id))

                                    rows_affected = c.rowcount
                                    conn.commit()

                                    c.execute("SELECT id, username, name FROM users WHERE id = ?", (user_id,))
                                    after_update = c.fetchone()

                                    conn.close()
                                    return True, f"User updated successfully. Rows affected: {rows_affected}", before_update, after_update
                                except sqlite3.Error as e:
                                    return False, f"Database error: {str(e)}", None, None
                                except Exception as e:
                                    return False, f"Error updating user: {str(e)}", None, None

                            success, message, before, after = update_user_with_path(user_id, name, email, phone,
                                                                                    department, role, join_date)

                            if success:
                                st.success(message)
                                st.write(f"Before update: {before}")
                                st.write(f"After update: {after}")

                                if reset_password:
                                    if not new_password.strip():
                                        st.error("Please enter a new password")
                                    else:

                                        def change_password_with_path(user_id, new_password):
                                            try:
                                                conn = sqlite3.connect(db_path)
                                                c = conn.cursor()

                                                hashed_password = hashlib.sha256(new_password.encode()).hexdigest()
                                                c.execute("UPDATE users SET password = ? WHERE id = ?",
                                                          (hashed_password, user_id))

                                                conn.commit()
                                                conn.close()
                                                return True, "Password changed successfully"
                                            except sqlite3.Error as e:
                                                return False, f"Password change failed: {str(e)}"

                                        pw_success, pw_message = change_password_with_path(user_id, new_password)
                                        if pw_success:
                                            st.success(pw_message)
                                        else:
                                            st.error(pw_message)

                                
                                if st.button("Reload Page"):
                                    st.rerun()
                            else:
                                st.error(message)
                else:
                    st.warning("Selected user not found")
            else:
                st.info("No users found to edit")
    
    elif menu == "Leave Management":
        st.header("Leave Management")
        tab1, tab2, tab3 = st.tabs(["Pending Approvals", "Leave History", "Add Leave for Employee"])
    
        with tab1:
            pending_leaves = get_pending_leave_requests()
            if not pending_leaves.empty:
                st.subheader("Pending Leave Requests")
                for _, leave in pending_leaves.iterrows():
                    with st.expander(f"{leave['name']} - {leave['start_date']} to {leave['end_date']}"):
                        st.markdown(f"**Type:** {leave['leave_type']}")
                        st.markdown(f"**Reason:** {leave['reason']}")
                        
                        col1, col2 = st.columns(2)
                        approval_note = col1.text_input("Approval Note", key=f"note_{leave['id']}")
                        
                        col2.markdown("##")
                        if col2.button("Approve", key=f"approve_{leave['id']}"):
                            update_leave_status(leave['id'], "Approved", approval_note)
                            st.success("Leave approved")
                            st.rerun()
                        if col2.button("Reject", key=f"reject_{leave['id']}"):
                            update_leave_status(leave['id'], "Rejected", approval_note)
                            st.success("Leave rejected")
                            st.rerun()
            else:
                st.info("No pending leave requests")

        with tab2:
            st.subheader("Leave Request History")
            # Filters
            col1, col2, col3 = st.columns(3)
            with col1:
                status_filter = st.selectbox("Filter by Status", ["All", "Approved", "Rejected", "Pending"])
            with col2:
                start_date_filter = st.date_input("Start Date", value=datetime.now() - timedelta(days=30))
            with col3:
                end_date_filter = st.date_input("End Date", value=datetime.now())
            
            #
            all_leaves = get_all_leave_requests(
                status_filter if status_filter != "All" else None,
                start_date_filter.strftime("%Y-%m-%d"),
                end_date_filter.strftime("%Y-%m-%d")
            )
            if not all_leaves.empty:
                
                search_query = st.text_input("Search by Employee Name")

                if search_query:
                    all_leaves = all_leaves[all_leaves['name'].str.contains(search_query, case=False)]
                
              
                for _, leave in all_leaves.iterrows():
                    status_color = {
                        "Approved": "🟢",
                        "Rejected": "🔴",
                        "Pending": "🟡"
                    }.get(leave['status'], "⚪")
                    
                    with st.expander(f"{status_color} {leave['name']} - {leave['start_date']} to {leave['end_date']}"):
                        col1, col2 = st.columns(2)
                        with col1:
                            st.markdown(f"**Employee:** {leave['name']}")
                            st.markdown(f"**Type:** {leave['leave_type']}")
                            st.markdown(f"**Reason:** {leave['reason']}")
                        with col2:
                            st.markdown(f"**Status:** {leave['status']}")
                            st.markdown(f"**Admin Note:** {leave['admin_approval'] or 'None'}")
                            st.markdown(f"**Verification:** {leave['employee_verification'] or 'Pending'}")
                        
                        st.markdown(f"**Request Date:** {leave['created_at']}")

                csv = all_leaves.to_csv(index=False)
                st.download_button(
                    label="Download Leave History",
                    data=csv,
                    file_name="leave_history.csv",
                    mime="text/csv"
                )
            else:
                st.info("No leave requests found for the selected filters")
        with tab3:
            st.subheader("Add Leave for Employee")

            users_df = get_all_users()
            if not users_df.empty:
                user_options = {row['name']: row['id'] for _, row in users_df.iterrows()}
                selected_name = st.selectbox("Select Employee", list(user_options.keys()))
                selected_user_id = user_options[selected_name]

                with st.form("admin_add_leave_form"):
                    col1, col2 = st.columns(2)
                    with col1:
                        start_date = st.date_input("Start Date")
                        leave_type = st.selectbox("Leave Type", ["Sick", "Vacation", "Personal", "Other"])
                    with col2:
                        end_date = st.date_input("End Date")
                        reason = st.text_area("Reason")

                    admin_note = st.text_input("Admin Note (Optional)")

                    if st.form_submit_button("Submit Leave"):
                        if start_date > end_date:
                            st.error("End date must be after start date")
                        else:
                            create_leave_request(
                                user_id=selected_user_id,
                                start_date=start_date.strftime("%Y-%m-%d"),
                                end_date=end_date.strftime("%Y-%m-%d"),
                                leave_type=leave_type,
                                reason=reason,
                                status="Approved", 
                                admin_approval=admin_note or "Leave added by admin"
                            )
                            st.success("Leave approved and added successfully")
                            st.rerun()
            else:
                st.warning("No users found in the system.")

    elif menu == "Attendance Management":
        st.header("Attendance Management")

        tab1, tab2, tab3 = st.tabs(["Today's Attendance", "View Attendance", "Attendance Summary"])
        # New Today's Attendance tab
        with tab1:
                st.markdown(f"<h3 style='color: #2d3748; margin-bottom: 20px;'>Today's Attendance - {datetime.now().strftime('%Y-%m-%d')}</h3>", 
               unsafe_allow_html=True)
                # Get today's attendance data
                todays_attendance = get_todays_attendance()
                
                if not todays_attendance.empty:
                    # Display statistics
                    present_count = len(todays_attendance)
                    checked_out_count = len(todays_attendance[todays_attendance['check_out'].notna()])
                    
                    col1, col2 = st.columns(2)
                    col1.metric("Employees Present Today", present_count)
                    #col2.metric("Employees Checked Out", checked_out_count)
                    
                    # Display the attendance data
                    st.dataframe(todays_attendance, use_container_width=True)
                    
                    # Download option
                    csv = todays_attendance.to_csv(index=False)
                    st.download_button(
                        label="Download Today's Attendance",
                        data=csv,
                        file_name=f"attendance_{datetime.now().strftime('%Y-%m-%d')}.csv",
                        mime="text/csv"
                    )
                else:
                    st.info("No attendance records found for today")

        with tab2:
            st.subheader("View Attendance Records")

            # Date Range
            col1, col2 = st.columns(2)
            with col1:
                start_date = st.date_input("From", value=datetime.now() - timedelta(days=7))
            with col2:
                end_date = st.date_input("To", value=datetime.now())

            # Filter
            filter_by = st.selectbox("Filter By", ["All Employees", "Department", "Individual"])
            attendance_df = None

            # Helper: Calculate working days excluding Sundays
            def get_working_days(start, end):
                return sum(1 for day in pd.date_range(start=start, end=end) if day.weekday() != 6)

            total_working_days = get_working_days(start_date, end_date)

            if filter_by == "All Employees":
                if start_date > end_date:
                    st.error("End date must be after start date")
                else:
                    attendance_df = get_all_attendance(start_date.strftime("%Y-%m-%d"), end_date.strftime("%Y-%m-%d"))

            elif filter_by == "Department":
                conn = sqlite3.connect('attendance_system.db')
                c = conn.cursor()
                c.execute("SELECT DISTINCT department FROM users WHERE department IS NOT NULL")
                departments = [dept[0] for dept in c.fetchall()]
                conn.close()

                selected_department = st.selectbox("Select Department", departments)

                if start_date > end_date:
                    st.error("End date must be after start date")
                else:
                    conn = sqlite3.connect('attendance_system.db')
                    query = """
                    SELECT u.name, a.date, a.check_in, a.check_out, a.work_hours, a.status, a.verification_method
                    FROM attendance a
                    JOIN users u ON a.user_id = u.id
                    WHERE a.date BETWEEN ? AND ? AND u.department = ?
                    ORDER BY a.date DESC, u.name
                    """
                    attendance_df = pd.read_sql_query(query, conn, params=(
                        start_date.strftime("%Y-%m-%d"),
                        end_date.strftime("%Y-%m-%d"),
                        selected_department
                    ))
                    conn.close()

            elif filter_by == "Individual":
                users_df = get_all_users()
                employees = users_df[users_df['role'] == 'employee']

                if not employees.empty:
                    # Create search interface
                    selected_employee = st.selectbox(
                        "Select Employee", 
                        employees['name'].tolist(),
                        key="employee_select"
                    )
                    
                    if st.button("🔍 Search", key="search_btn"):
                        try:
                            # Debug: Show selected employee name
                            st.write(f"Searching for: {selected_employee}")
                            
                            # Get employee ID
                            selected_user_id = int(employees[employees['name'] == selected_employee]['id'].iloc[0])
                            st.write(f"Employee ID: {selected_user_id}")
                            
                            if start_date > end_date:
                                st.error("End date must be after start date")
                            else:
                                # Format dates for SQL query
                                start_str = start_date.strftime("%Y-%m-%d")
                                end_str = end_date.strftime("%Y-%m-%d")
                                st.write(f"Date range: {start_str} to {end_str}")
                                
                                # Get attendance data with debug info
                                conn = sqlite3.connect('attendance_system.db')
                                query = """
                                SELECT date, check_in, check_out, work_hours, status, verification_method
                                FROM attendance 
                                WHERE user_id = ? AND date BETWEEN ? AND ?
                                ORDER BY date DESC
                                """
                                attendance_df = pd.read_sql_query(query, conn, params=(selected_user_id, start_str, end_str))
                                conn.close()
                                
                                st.write(f"Found {len(attendance_df)} records")  # Debug count

                                if not attendance_df.empty:
                                    # Calculate metrics
                                    present_days = attendance_df.shape[0]
                                    total_work_hours = attendance_df['work_hours'].sum()
                                    
                                    # Format work hours to HH:MM
                                    attendance_df['work_hours'] = attendance_df['work_hours'].apply(
                                        lambda x: f"{int(x)}:{int((x % 1)*60):02d}" if pd.notnull(x) else "N/A"
                                    )

                                    # Display header and metrics
                                    st.success(f"Attendance records found for {selected_employee}")
                                    st.markdown(f"**Date Range:** {start_str} to {end_str}")
                                    
                                    col1, col2, col3 = st.columns(3)
                                    col1.metric("Working Days", (end_date - start_date).days + 1)
                                    col2.metric("Present Days", present_days)
                                    col3.metric("Total Work Hours", f"{total_work_hours:.2f}")
                                    
                                    # Display detailed attendance table
                                    st.dataframe(
                                        attendance_df.rename(columns={
                                            'date': 'Date',
                                            'check_in': 'Check-In Time',
                                            'check_out': 'Check-Out Time',
                                            'work_hours': 'Work Hours',
                                            'status': 'Status',
                                            'verification_method': 'Verification Method'
                                        }),
                                        use_container_width=True,
                                        hide_index=True
                                    )

                                    # Download button
                                    csv = attendance_df.to_csv(index=False)
                                    st.download_button(
                                        label="📥 Download Attendance",
                                        data=csv,
                                        file_name=f"{selected_employee.replace(' ', '_')}_attendance_{start_str}_to_{end_str}.csv",
                                        mime="text/csv"
                                    )
                                else:
                                    st.warning(f"No attendance records found for {selected_employee} between {start_str} and {end_str}")
                                    # Debug: Show all dates for this employee
                                    conn = sqlite3.connect('attendance_system.db')
                                    all_dates_query = """
                                    SELECT date FROM attendance 
                                    WHERE user_id = ?
                                    ORDER BY date DESC
                                    LIMIT 10
                                    """
                                    all_dates = pd.read_sql_query(all_dates_query, conn, params=(selected_user_id,))
                                    conn.close()
                                    if not all_dates.empty:
                                        st.write(f"Most recent dates for this employee: {all_dates['date'].tolist()}")
                                    else:
                                        st.write("No attendance records at all for this employee")
                        except Exception as e:
                            st.error(f"Error searching attendance: {str(e)}")
                            st.write("Please check the logs for more details")
                else:
                    st.info("No employees found in the system")


            # If general filter: show data and summary
            if attendance_df is not None and filter_by != "Individual":
                if not attendance_df.empty:
                    present_days = attendance_df['date'].nunique()
                    total_work_hours = attendance_df['work_hours'].sum() if 'work_hours' in attendance_df.columns else 0

                    # Summary
                    col1, col2, col3 = st.columns(3)
                    with col1:
                        st.metric("Working Days (Excl. Sundays)", total_working_days)
                    with col2:
                        st.metric("Total Records", attendance_df.shape[0])
                    with col3:
                        st.metric("Total Work Hours", f"{total_work_hours:.2f}")

                    # Table
                    st.dataframe(attendance_df, use_container_width=True)
                    csv = attendance_df.to_csv(index=False)
                    st.download_button(
                        label="Download as CSV",
                        data=csv,
                        file_name=f"attendance_{start_date}_to_{end_date}.csv",
                        mime="text/csv"
                    )
                else:
                    st.info("No attendance records found for selected criteria")

        with tab3:
            st.subheader("Attendance Summary")
            # ... (rest of the Attendance Summary tab remains the same)
            with tab3:
                # Select month and year
                col1, col2 = st.columns(2)

                with col1:
                    selected_month = st.selectbox("Month", [
                        "January", "February", "March", "April", "May", "June",
                        "July", "August", "September", "October", "November", "December"
                    ])
                    month_num = {
                        "January": 1, "February": 2, "March": 3, "April": 4,
                        "May": 5, "June": 6, "July": 7, "August": 8,
                        "September": 9, "October": 10, "November": 11, "December": 12
                    }[selected_month]

                with col2:
                    current_year = datetime.now().year
                    selected_year = st.selectbox("Year", range(current_year - 2, current_year + 1), index=2)

                # Generate report
                attendance_summary = generate_attendance_report(month_num, selected_year)

                if not attendance_summary.empty:
                    st.dataframe(attendance_summary, use_container_width=True)

                    # Download option
                    csv = attendance_summary.to_csv(index=False)
                    st.download_button(
                        label="Download Summary",
                        data=csv,
                        file_name=f"attendance_summary_{selected_month}_{selected_year}.csv",
                        mime="text/csv"
                    )
                else:
                    st.info("No attendance data found for the selected month and year")
            
    elif menu == "Reports":
        st.subheader("All Daily Reports")

        
        col1, col2, col3, col4 = st.columns(4)

        with col1:
            start_date = st.date_input("From", value=datetime.now() - timedelta(days=30), key="reports_start")
        with col2:
            end_date = st.date_input("To", value=datetime.now(), key="reports_end")
        with col3:

            conn = sqlite3.connect("attendance_system.db")
            c = conn.cursor()

            
            c.execute("SELECT project FROM daily_reports")
            projects = c.fetchall()
            conn.close()

           
            project_list = sorted(set(
                p.strip() for row in projects for p in row[0].split(",") if p.strip()
            ))

           
            if "General" not in project_list:
                project_list.append("General")

           
            selected_projects = st.multiselect("Filter by Project(s)", options=project_list)
        with col4:
            search_emp = st.text_input("Search by Employee Name or ID")

        
        if start_date > end_date:
            st.error("⚠️ End date must be after start date")
        else:
            
            conn = sqlite3.connect("attendance_system.db")
            query = (
                "SELECT u.id AS EmployeeID, u.name AS Employee, r.date AS Date, r.project AS Project,"
                " r.activities AS Activities, r.achievements AS Achievements, r.challenges AS Challenges"
                " FROM daily_reports r"
                " JOIN users u ON r.user_id = u.id"
                " WHERE r.date BETWEEN ? AND ?"
            )
            params = [start_date.strftime("%Y-%m-%d"), end_date.strftime("%Y-%m-%d")]

            
            if selected_projects:
                like_conditions = " OR ".join(["r.project LIKE ?"] * len(selected_projects))
                query += f" AND ({like_conditions})"
                params.extend([f"%{p}%" for p in selected_projects])


           
            if search_emp:
                
                query += " AND (u.name LIKE ? OR u.id = ?)"
                params.append(f"%{search_emp}%")
                
                try:
                    params.append(int(search_emp))
                except ValueError:
                    params.append(-1)

            query += " ORDER BY r.date DESC, u.name"

            
            df = pd.read_sql_query(query, conn, params=params)
            conn.close()

            
            if df.empty:
                st.info("No reports found for the selected filters.")
            else:
                st.dataframe(df, use_container_width=True)

              
                csv = df.to_csv(index=False)
                st.download_button(
                    label="📥 Download Reports as CSV",
                    data=csv,
                    file_name=f"daily_reports_{start_date}_to_{end_date}.csv",
                    mime="text/csv"
                )

    elif menu == "Project Reports":
        st.header("Project Reports")
        projects = get_projects()
        
        if projects:
            # Update columns to match all 12 fields from the database
            df = pd.DataFrame(projects, columns=[
                "ID", "Name", "Description", "Link", "Members", "Deadline", 
                "Total Hours", "Status", "Client Name", "Client Phone", 
                "Client Email", "Extra Details"
            ])
            
            st.dataframe(df, use_container_width=True)
            
            # Update CSV download to match the columns
            csv = df.to_csv(index=False)
            st.download_button(
                label="Download Project Report",
                data=csv,
                file_name="sales_project_report.csv",
                mime="text/csv"
            )
        else:
            st.info("No project reports found.")
    
            
    
    elif menu == "Salary Management":
        st.header("Salary Management")
        
        tab1, tab2 = st.tabs(["Process Salary", "Salary History"])
        
        with tab1:
            st.subheader("Process Monthly Salary")
            
            col1, col2 = st.columns(2)
            
            with col1:
                selected_month = st.selectbox("Month", [
                    "January", "February", "March", "April", "May", "June",
                    "July", "August", "September", "October", "November", "December"
                ], key="process_month")
                month_num = {
                    "January": 1, "February": 2, "March": 3, "April": 4,
                    "May": 5, "June": 6, "July": 7, "August": 8,
                    "September": 9, "October": 10, "November": 11, "December": 12
                }[selected_month]
            
            with col2:
                current_year = datetime.now().year
                selected_year = st.selectbox("Year", range(current_year-2, current_year+1), index=2, key="process_year")
            
            users_df = get_all_users()
            employees = users_df[users_df['role'] == 'employee']
            
            search_name = st.text_input("Search Employee by Name").lower()
            
            department_filter = st.selectbox("Filter by Department", ["All"] + list(employees['department'].unique()))

            if department_filter != "All":
                filtered_employees = employees[
                    employees['name'].str.lower().str.contains(search_name) &
                    (employees['department'] == department_filter)
                ]
            else:
                filtered_employees = employees[employees['name'].str.lower().str.contains(search_name)]

            if not filtered_employees.empty:
                
                salary_data = []
                st.markdown("### 🧾 Employee Salary List")

                for idx, employee in filtered_employees.iterrows():
                    employee_id = employee['id']
                    employee_name = employee['name']

                    summary = get_monthly_attendance_summary(employee_id, month_num, selected_year)
                    present_days = summary.get('present_days', 0)
                    work_hours = summary.get('work_hours', 0)

                    conn = sqlite3.connect('attendance_system.db')
                    c = conn.cursor()
                    c.execute("""
                        SELECT base_salary, bonus, deductions, total_salary, payment_status
                        FROM salary WHERE user_id = ? AND month = ? AND year = ?
                    """, (employee_id, month_num, selected_year))
                    existing_salary = c.fetchone()
                    conn.close()
                    
                    with st.expander(f"💼 {employee_name}"):
                        col1, col2, col3 = st.columns(3)

                        with col1:
                            st.write("**Present Days:**", present_days)
                            base_salary = st.number_input("Base Salary (₹)", value=existing_salary[0] if existing_salary else 0.0,
                                                    step=100.0, key=f"base_{employee_id}")

                        with col2:
                            st.write("**Work Hours:**", f"{work_hours:.2f}")
                            bonus = st.number_input("Bonus (₹)", value=existing_salary[1] if existing_salary else 0.0,
                                                step=10.0, key=f"bonus_{employee_id}")

                        with col3:
                            st.write("**Total Salary (₹):**")
                            deductions = st.number_input("Deductions (₹)", value=existing_salary[2] if existing_salary else 0.0,
                                                        step=10.0, key=f"deduct_{employee_id}")
                            total_salary = base_salary + bonus - deductions
                            st.success(f"₹ {total_salary:.2f}")

                        payment_status = st.selectbox("Payment Status", ["Pending", "Paid"],
                                                    index=0 if not existing_salary or existing_salary[4] == "Pending" else 1,
                                                    key=f"status_{employee_id}")

                        if st.button(f"Save for {employee_name}", key=f"save_{employee_id}"):
                            save_salary_info(employee_id, month_num, selected_year, base_salary, bonus, deductions, total_salary, payment_status)
                            st.success(f"Saved for {employee_name}")

                        salary_data.append({
                            'Employee ID': employee_id,
                            'Employee Name': employee_name,
                            'Base Salary': base_salary,
                            'Bonus': bonus,
                            'Deductions': deductions,
                            'Total Salary': total_salary,
                            'Status': payment_status
                        })
            else:
                st.warning("No matching employees found.")
        with tab2:

            st.subheader("Salary History")

            col1, col2 = st.columns(2)  

            with col1:
                month_filter = st.selectbox("Month", ["All"] + [
                    "January", "February", "March", "April", "May", "June",
                    "July", "August", "September", "October", "November", "December"
                ], key="month_select")

            with col2:
                current_year = datetime.now().year
                year_filter = st.selectbox("Year", ["All"] + list(range(current_year - 2, current_year + 1)), key="year_select")

            salary_history = get_salary_history(
                month_filter if month_filter != "All" else None,
                year_filter if year_filter != "All" else None
            )

            search_name = st.text_input("Search Employee by Name", key="search_name")

            search_button = st.button("Search", key="search_button")

            if not salary_history.empty:
                salary_history['year'] = salary_history['year'].astype(str)

                if search_button and search_name:
                    salary_history = salary_history[salary_history['employee_name'].str.contains(search_name, case=False)]

                if not salary_history.empty:
                    st.dataframe(salary_history, use_container_width=True)

                    csv = salary_history.to_csv(index=False)
                    st.download_button(
                        label="Download Salary History",
                        data=csv,
                        file_name=f"salary_history_{month_filter}_{year_filter}.csv",
                        mime="text/csv"
                    )
                else:
                    st.warning("No records match your search.")
            else:
                st.info("No salary history found for the selected filters")

    elif menu == "Settings":
        st.header("System Settings")

        tab1, tab2 = st.tabs(["General Settings", "Backup/Restore"])

        with tab1:
            st.header(" Post Announcement")
            title = st.text_input("Title")
            message = st.text_area("Announcement Message")

            if st.button("Post Announcement"): 
                if title and message:
                    conn = sqlite3.connect('attendance_system.db')
                    c = conn.cursor()
                    c.execute("INSERT INTO announcements (title, message, created_at) VALUES (?, ?, ?)",
                            (title, message, datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
                    conn.commit()
                    conn.close()
                    st.success("✅ Announcement posted!")
                else:
                    st.warning("Please fill out both fields.")

        with tab2:
            st.subheader("Backup and Restore")

            if st.button("Backup Database"):
 
                backup_file = f"attendance_system_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db"

                shutil.copy('attendance_system.db', backup_file)

                with open(backup_file, 'rb') as f:
                    backup_data = f.read()

                st.download_button(
                    label="Download Backup File",
                    data=backup_data,
                    file_name=backup_file,
                    mime="application/octet-stream"
                )

                os.remove(backup_file)

                st.success("Database backup created successfully!")

            st.write("#### Restore Database")
            uploaded_file = st.file_uploader("Upload backup file", type=["db"])

            if uploaded_file is not None:
                
                st.warning("Restoring from backup will overwrite the current database. This cannot be undone.")
                if st.button("Restore Database"):
                    
                    with open('temp_restore.db', 'wb') as f:
                        f.write(uploaded_file.getbuffer())

                    try:
                        conn = sqlite3.connect('temp_restore.db')
                        c = conn.cursor()
                        c.execute("SELECT name FROM sqlite_master WHERE type='table'")
                        tables = c.fetchall()
                        conn.close()

                        required_tables = ['users', 'attendance', 'daily_reports', 'salary']
                        table_names = [table[0] for table in tables]
                        missing_tables = [table for table in required_tables if table not in table_names]

                        if missing_tables:
                            st.error(f"Invalid database backup. Missing tables: {', '.join(missing_tables)}")
                            os.remove('temp_restore.db')
                        else:
                            
                            backup_file = f"attendance_system_auto_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db"
                            shutil.copy('attendance_system.db', backup_file)

                            os.remove('attendance_system.db')
                            shutil.move('temp_restore.db', 'attendance_system.db')

                            st.success("Database restored successfully! Please restart the application.")
                            st.balloons()
                    except Exception as e:
                        st.error(f"Error restoring database: {str(e)}")
    
                        if os.path.exists('temp_restore.db'):
                            os.remove('temp_restore.db')
def sales_dashboard():
    st.title("Sales Dashboard")
    if st.session_state.get("user_role") == "admin":
        if st.button("🔙 Back to Admin Dashboard"):
            st.session_state.is_sales_login = False
            st.rerun()
    else:
        st.info("You are logged in as a Sales user.")
    
    menu = st.sidebar.selectbox(
        "Menu",
        ["Home", "All Projects", "Add Project", "Edit Project", "Allocated Project"],
        index=["Home", "All Projects", "Add Project", "Edit Project", "Allocated Project"].index(
            st.session_state.get("sales_tab", "Home")
        )
    )
    st.session_state.sales_tab = menu

    if st.sidebar.button("Logout"):
        logout()
    

    if menu == "Home":
        st.header("Dashboard Overview")
        projects = get_projects()
        total = len(projects)
        ongoing = len([p for p in projects if p[7] == "Ongoing"])
        finished = len([p for p in projects if p[7] == "Finished"])
        not_assigned = len([p for p in projects if p[7] == "Not Assigned"])

        col1, col2, col3, col4 = st.columns(4)
        col1.metric("Total Projects", total)
        col2.metric("Ongoing", ongoing)
        col3.metric("Finished", finished)
        col4.metric("Not Assigned", not_assigned)

    elif menu == "All Projects":
        display_project_directory("All Projects")


    elif menu == "Add Project":

        st.header("Add New Project")

        with st.form("add_project_form"):

            name = st.text_input("Project Name")

            desc = st.text_area("Description")

            client_name = st.text_input("Client Name")

            client_phone = st.text_input("Client Phone Number", max_chars=10)

            client_email = st.text_input("Client Email ID")

            extra_details = st.text_area("Client Details")

            link = st.text_input("Project Link")

            all_usernames = get_all_users()

            all_usernames = [row['username'] for _, row in all_usernames.iterrows() if row['role'] == "employee"]

            members = st.multiselect("Assign Members", all_usernames)

            deadline = st.date_input("Deadline")

            hours = st.number_input("Total Hours", min_value=0.0)

            submit = st.form_submit_button("Add Project")

        if submit:
            members_str = ", ".join(members)

            add_project(name, desc, link, members_str, deadline.strftime("%Y-%m-%d"), hours,

                        client_name, client_phone, client_email, extra_details)

            st.success("✅ Project added successfully!")


    elif menu == "Edit Project":

        st.header("Edit Project")

        projects = get_projects()

        project_dict = {f"{p[0]} - {p[1]}": p for p in projects}

        selected = st.selectbox("Select Project", list(project_dict.keys()))

        project = project_dict[selected]

        with st.form("edit_project_form"):
            name = st.text_input("Project Name", value=project[1])
            desc = st.text_area("Description", value=project[2])
            link = st.text_input("Project Link", value=project[3])
            all_employees = get_employees()
            default_members = [m.strip() for m in project[4].split(",") if m.strip() in all_employees]
            members = st.multiselect("Assign Members", all_employees, default=default_members)
            deadline = st.date_input("Deadline", value=datetime.strptime(project[5], "%Y-%m-%d"))
            hours = st.number_input("Total Hours", value=project[6])

            status_options = ["Not Assigned", "Ongoing", "Finished", "Active"]
            default_status = project[7] if project[7] in status_options else "Not Assigned"
            status = st.selectbox("Status", status_options, index=status_options.index(default_status))

            client_name = st.text_input("Client Name", value=project[8] if project[8] else "")
            client_phone = st.text_input("Client Phone Number", value=project[9] if project[9] else "", max_chars=10)

            if client_phone and not client_phone.isdigit():
                st.warning("Only numeric digits allowed.")
                client_phone = ''.join(filter(str.isdigit, client_phone))[:10]

            client_email = st.text_input("Client Email", value=project[10] if project[10] else "")
            extra_details = st.text_area("Extra Details", value=project[11] if project[11] else "")

            update = st.form_submit_button("Update Project")  # <-- No reassignment to False

            if update:
                if not client_phone.isdigit() or len(client_phone) != 10:
                    st.error("❌ Client Phone Number must be exactly 10 digits.")
                elif not client_email.endswith("@gmail.com"):
                    st.error("❌ Client Email must be a valid Gmail address.")
                else:
                    update_project(
                        project_id=project[0],
                        name=name,
                        desc=desc,
                        link=link,
                        members=members,
                        deadline=deadline.strftime("%Y-%m-%d"),
                        hours=hours,
                        status=status,
                        client_name=client_name,
                        client_phone=client_phone,
                        client_email=client_email,
                        extra_details=extra_details,
                    )
                    st.success("✅ Project updated successfully!")



    elif menu == "Allocated Project":

        st.header("Allocated Projects")

        # Fetch all projects

        projects = get_projects()

        if not projects:

            st.info("No allocated projects available.")

        else:

            df = pd.DataFrame(projects, columns=[

                "ID", "Name", "Description", "Link", "Members", "Deadline", "Total Hours", "Status",

                "Client Name", "Client Phone", "Client Email", "Extra Details"

            ])

            # Filter only those with assigned members

            df = df[df["Members"].str.strip() != ""]

            # Display projects

            for _, row in df.iterrows():
                with st.expander(f"{row['Name']}"):
                    st.markdown(f"**Client Name:** {row['Client Name']}")

                    st.markdown(f"**Client Phone:** {row['Client Phone']}")

                    st.markdown(f"**Client Email:** {row['Client Email']}")

                    st.markdown(f"**Extra Details:** {row['Extra Details']}")

                    st.markdown(f"**Assigned Members:** {row['Members']}")

                    st.markdown(f"**Deadline:** {row['Deadline']}")

                    st.markdown(f"**Total Hours:** {row['Total Hours']}")

                    st.markdown(f"**Status:** {row['Status']}")

                    st.markdown(f"[🔗 Project Link]({row['Link']})")

        update = False

        if update:
            update_project(project[0], name, desc, link, members, deadline.strftime("%Y-%m-%d"), hours, status)

            st.success("✅ Project updated successfully!")


def main():
    
    if 'authenticated' not in st.session_state:
        st.session_state.authenticated = False

    if 'user_id' not in st.session_state:
        st.session_state.user_id = None

    if 'user_role' not in st.session_state:
        st.session_state.user_role = None

    if 'face_encodings_db' not in st.session_state:
        st.session_state.face_encodings_db = load_face_encodings()

    if 'is_sales_login' not in st.session_state:
        st.session_state.is_sales_login = False


    # Initialize database
    init_db()

    if not st.session_state.authenticated:
        login_page()
    else:
        if st.session_state.get('is_sales_login', False):
            sales_dashboard()
        else:
            if st.session_state.user_role == 'admin':
                admin_dashboard()
            else:
                employee_dashboard()


def get_user_role(user_id):
    conn = sqlite3.connect('attendance_system.db')
    c = conn.cursor()
    c.execute("SELECT role FROM users WHERE id = ?", (user_id,))
    result = c.fetchone()
    conn.close()
    return result[0] if result else None

def delete_user(user_id):
    conn = sqlite3.connect('attendance_system.db')
    c = conn.cursor()
    c.execute("DELETE FROM users WHERE id = ?", (user_id,))
    conn.commit()
    conn.close()
def check_liveness(image_array):
    
    eye_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + 'haarcascade_eye.xml')
    gray = cv2.cvtColor(image_array, cv2.COLOR_BGR2GRAY)
    eyes = eye_cascade.detectMultiScale(gray, 1.1, 4)
    return len(eyes) >= 2 
def detect_liveness(image_array):
    """Enhanced liveness detection using head orientation and eye blink"""
    img_rgb = cv2.cvtColor(image_array, cv2.COLOR_BGR2RGB)
    results = face_mesh.process(img_rgb)
    
    if results.multi_face_landmarks:
        for face_landmarks in results.multi_face_landmarks:
            
            direction = detect_direction(face_landmarks.landmark, image_array.shape[1], image_array.shape[0])
            
            
            left_eye = [face_landmarks.landmark[i] for i in [33, 160, 158, 133]]
            right_eye = [face_landmarks.landmark[i] for i in [362, 385, 387, 263]]
            
            if is_blinking(left_eye) or is_blinking(right_eye):
                return True, direction
            
    return False, None

def biometric_registration():
    st.subheader("Face Registration")
  
    if st.button("Start Face Registration"):
        success = register_face(st.session_state.user_id)
        if success:
            st.success("Registered face with multiple angles!")
        else:
            st.error("Registration failed")
def detect_direction(landmarks, width, height):
    """Detect head orientation using facial landmarks"""
    nose_tip = landmarks[1]
    left_eye = landmarks[33]
    right_eye = landmarks[263]

    nose_x = nose_tip.x * width
    eye_center_x = (left_eye.x + right_eye.x) * width / 2
    
    if nose_x < eye_center_x - 30:
        return 'left'
    elif nose_x > eye_center_x + 30:
        return 'right'
    return 'center'
def face_login():
    st.write("Look at the camera")
    img_file = st.camera_input("Real-time verification")
    
    if img_file:
        image = Image.open(img_file)
        image_array = np.array(image)
        
        if verify_face(image_array, st.session_state.user_id):
            st.success("Biometric verification passed!")
        else:
            st.error("Verification failed")
def is_blinking(eye_points):
    """Detect eye blink using EAR (Eye Aspect Ratio)"""
    vertical = [
        math.dist((eye_points[0].x, eye_points[0].y), 
        (eye_points[1].x, eye_points[1].y)),
        math.dist((eye_points[2].x, eye_points[2].y), 
        (eye_points[3].x, eye_points[3].y))
    ]
    horizontal = math.dist(
        (eye_points[0].x, eye_points[0].y),
        (eye_points[3].x, eye_points[3].y)
    )
    ear = (vertical[0] + vertical[1]) / (2.0 * horizontal)
    return ear < 0.2  # Threshold for blink detection
def render_custom_menu(menu_items, current_menu):
    # Hamburger icon
    components.html("""
    <div class="hamburger" onclick="toggleMenu()">
        <div></div>
        <div></div>
        <div></div>
    </div>
    """, height=80)
    
    # Menu items
    menu_html = """
    <div class="sidebar-menu">
        <h3>Menu</h3>
        <div style="margin-top: 20px;">
    """
    
    for item in menu_items:
        active_class = "active" if item == current_menu else ""
        menu_html += f"""
        <div class="menu-item {active_class}" 
             onclick="handleMenuItemClick('{item}')">
            {item}
        </div>
        """
    
    menu_html += "</div></div>"
    
    components.html(menu_html, height=600)

def display_user_directory(title, role_filter):
    st.header(title)


    users_df = get_all_users()
    users = users_df[users_df['role'] == role_filter]

    if not users.empty:
        col1, col2 = st.columns(2)

        with col1:
            department_filter = st.selectbox(
                "Filter by Department",
                ["All Departments"] + sorted(users['department'].dropna().unique().tolist()),
                key=f"{role_filter}_dept"
            )
        with col2:
            search_query = st.text_input("Search by Name", key=f"{role_filter}_search")

        if department_filter != "All Departments":
            users = users[users['department'] == department_filter]
        if search_query:
            users = users[users['name'].str.contains(search_query, case=False)]

        st.subheader("User Cards")
        cols = st.columns(3)
        for idx, (_, user) in enumerate(users.iterrows()):
            with cols[idx % 3]:
                with st.container():
                    conn = sqlite3.connect('attendance_system.db')
                    c = conn.cursor()
                    c.execute("SELECT photo FROM users WHERE id = ?", (user['id'],))
                    photo_blob = c.fetchone()[0]
                    conn.close()

                    if photo_blob:
                        st.image(photo_blob, width=150)
                    else:
                        st.image("https://via.placeholder.com/150?text=No+Photo", width=150)

                    st.markdown(f"**{user['name']}**")
                    st.markdown(f"*{user['department']}*")

                    with st.expander("View Details"):
                        st.markdown(f"**Email:** {user['email']}")
                        st.markdown(f"**Phone:** {user['phone']}")
                        st.markdown(f"**Join Date:** {user['join_date']}")
                        
                        if user.get('professional_skills'):
                            st.markdown("---")
                            st.markdown("**Professional Skills:**")
                            skills_list = [skill.strip() for skill in user['professional_skills'].split(',') if skill.strip()]
                            for skill in skills_list:
                                st.markdown(f"- {skill}")
                        else:
                            st.markdown("**Professional Skills:** Not specified")

                    salary_df = get_user_salary(user['id'])
                    if not salary_df.empty:
                        st.markdown("---")
                        st.markdown("**Salary Info**")
                        st.markdown(f"**Total Salary:** ₹{latest_salary['total_salary']:.2f}")
                        st.markdown(f"**Status:** {latest_salary['payment_status']}")


                        salary_df = get_user_salary(user['id'])
                        if not salary_df.empty:
                            latest_salary = salary_df.iloc[0]
                            st.markdown("---")
                            st.markdown("**Salary Info**")
                            st.markdown(f"**Total Salary:** ₹{latest_salary['total_salary']:.2f}")
                            st.markdown(f"**Status:** {latest_salary['payment_status']}")
    else:
        st.info("No users found.")

def create_leave_request(user_id, start_date, end_date, leave_type, reason, status="Pending", admin_approval=""):
    try:
        conn = sqlite3.connect('attendance_system.db')
        cursor = conn.cursor()
        cursor.execute("""
            INSERT INTO leave_requests (user_id, start_date, end_date, leave_type, reason, status, admin_approval)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (user_id, start_date, end_date, leave_type, reason, status, admin_approval))
        conn.commit()
        conn.close()
    except Exception as e:
        st.error(f"Error creating leave request: {str(e)}")

def get_user_info(user_id):
    conn = sqlite3.connect('attendance_system.db')
    c = conn.cursor()
    
    c.execute("""
    SELECT username, name, email, phone, department, join_date, role, photo, professional_skills
    FROM users WHERE id = ?
    """, (user_id,))
    
    result = c.fetchone()
    conn.close()
    
    if result:
        return {
            'username': result[0],
            'name': result[1],
            'email': result[2],
            'phone': result[3],
            'department': result[4],
            'join_date': result[5],
            'role': result[6],
            'photo': result[7],
            'professional_skills': result[8] 
        }
    return None

def get_all_users():
    try:
        conn = sqlite3.connect('attendance_system.db')
        query = """
        SELECT id, username, name, email, phone, department, join_date, role, photo, professional_skills
        FROM users
        ORDER BY name
        """
        users_df = pd.read_sql_query(query, conn)
        conn.close()
        return users_df
    except Exception as e:
        st.error(f"Error fetching users: {str(e)}")
        return pd.DataFrame()

def get_employee_leave_requests(user_id):
    conn = sqlite3.connect('attendance_system.db')
    query = """
    SELECT id, start_date, end_date, leave_type, reason, status, admin_approval, employee_verification
    FROM leave_requests
    WHERE user_id = ?
    ORDER BY created_at DESC
    """
    df = pd.read_sql_query(query, conn, params=(user_id,))
    conn.close()
    return df

def get_all_leave_requests_by_user(user_id):
    conn = sqlite3.connect('attendance_system.db')
    query = """
        SELECT start_date, end_date, leave_type, status
        FROM leave_requests
        WHERE user_id = ?
    """
    df = pd.read_sql_query(query, conn, params=(user_id,))
    conn.close()
    return df


def get_pending_leave_requests():
    conn = sqlite3.connect('attendance_system.db')
    query = """
    SELECT 
        lr.id,
        u.name,
        lr.start_date,
        lr.end_date,
        lr.leave_type,
        lr.reason,
        lr.created_at
    FROM leave_requests lr
    JOIN users u ON lr.user_id = u.id
    WHERE lr.status = 'Pending'
    ORDER BY lr.created_at DESC
    """
    df = pd.read_sql_query(query, conn)
    conn.close()
    return df

def update_leave_status(leave_id, status, approval_note):
    conn = sqlite3.connect('attendance_system.db')
    c = conn.cursor()
    c.execute("""
    UPDATE leave_requests 
    SET status = ?, admin_approval = ?
    WHERE id = ?
    """, (status, approval_note, leave_id))
    conn.commit()
    conn.close()

def verify_leave_request(leave_id):
    conn = sqlite3.connect('attendance_system.db')
    c = conn.cursor()
    c.execute("""
    UPDATE leave_requests 
    SET employee_verification = 'Verified'
    WHERE id = ?
    """, (leave_id,))
    conn.commit()
    conn.close()

def get_employees():
    conn = sqlite3.connect('attendance_system.db')
    c = conn.cursor()
    c.execute("SELECT name FROM users WHERE role='employee'")
    employees = [row[0] for row in c.fetchall()]
    conn.close()
    return employees

def add_project(name, desc, link, members, deadline, hours, client_name, client_phone, client_email, extra_details):
    conn = sqlite3.connect('attendance_system.db')
    c = conn.cursor()
    members_str = ", ".join(members)
    status = "Not Assigned" if not members else "Ongoing"
    c.execute("""
        INSERT INTO projects (
            name, description, link, members, deadline, total_hours, status,
            client_name, client_phone, client_email, extra_details
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (name, desc, link, members, deadline, hours, "Not Assigned",
          client_name, client_phone, client_email, extra_details))

    conn.commit()
    conn.close()

def get_projects():
    conn = sqlite3.connect('attendance_system.db')
    c = conn.cursor()
    c.execute("SELECT * FROM projects")
    rows = c.fetchall()
    conn.close()
    return rows

def update_project(project_id, name, desc, link, members, deadline, hours, status,
                   client_name, client_phone, client_email, extra_details):
    conn = sqlite3.connect('attendance_system.db')
    c = conn.cursor()
    members_str = ", ".join(members)

    c.execute("""
        UPDATE projects
        SET name=?, description=?, link=?, members=?, deadline=?, total_hours=?, status=?,
            client_name=?, client_phone=?, client_email=?, extra_details=?
        WHERE id=?
    """, (
        name, desc, link, members_str, deadline, hours, status,
        client_name, client_phone, client_email, extra_details, project_id
    ))

    conn.commit()
    conn.close()

def display_project_directory(title):
    st.header(title)

    projects = get_projects()
    if not projects:
        st.info("No projects available.")
        return

    df = pd.DataFrame(projects, columns=[
        "ID", "Name", "Description", "Link", "Members", "Deadline", "Total Hours", "Status",
        "Client Name", "Client Phone", "Client Email", "Extra Details"
    ])

    col1, col2 = st.columns(2)
    with col1:
        status_filter = st.selectbox("Filter by Status", ["All"] + sorted(df["Status"].unique().tolist()))
    with col2:
        search_query = st.text_input("Search by Project Name")

    if status_filter != "All":
        df = df[df["Status"] == status_filter]
    if search_query:
        df = df[df["Name"].str.contains(search_query, case=False)]

    st.subheader("Projects")
    cols = st.columns(3)
    for idx, row in df.iterrows():
        with cols[idx % 3]:
            with st.container():
                st.markdown(f"""
                    <div style='
                        background-color: #f8f9fa;
                        padding: 15px;
                        border-radius: 10px;
                        box-shadow: 0 0 10px rgba(0,0,0,0.05);
                        margin-bottom: 15px;
                    '>
                        <h5>{row.get('Name', 'N/A')}</h5>
                        <p><strong>Members:</strong> {row.get('Members', 'N/A')}</p>
                        <p><strong>Deadline:</strong> {row.get('Deadline', 'N/A')}</p>
                        <p><strong>Status:</strong> 
                            <span style='color: {"green" if row.get("Status")=="Finished" else "orange" if row.get("Status")=="Ongoing" else "red"}'>
                                {row.get("Status", "N/A")}
                            </span>
                        </p>
                        <a href="{row.get('Link', '#')}" target="_blank">🔗 Project Link</a>
                    </div>
                """, unsafe_allow_html=True)

def add_user(username, password, role, name, email, phone, department, join_date, photo_bytes=None):
    conn = sqlite3.connect('attendance_system.db')
    c = conn.cursor()
    try:
        
        if not phone.isdigit() or len(phone) != 10:
            return False, None, "Invalid phone number format (must be 10 digits)"
        
        hashed_password = hashlib.sha256(password.encode()).hexdigest()
        formatted_join_date = join_date.strftime("%Y-%m-%d") if isinstance(join_date, datetime) else join_date
        
        c.execute("""
            INSERT INTO users (username, password, role, name, email, phone, department, join_date, photo)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (username, hashed_password, role, name, email, phone, department, formatted_join_date, photo_bytes))
        
        user_id = c.lastrowid
        conn.commit()
        conn.close()
        return True, user_id, "User added successfully"
    except sqlite3.IntegrityError:
        conn.close()
        return False, None, "Username already exists"
    except Exception as e:
        conn.close()
        return False, None, f"Error adding user: {str(e)}"
def get_all_leave_requests(status=None, start_date=None, end_date=None):
    conn = sqlite3.connect('attendance_system.db')
    query = """
    SELECT 
        lr.id,
        u.name,
        lr.start_date,
        lr.end_date,
        lr.leave_type,
        lr.reason,
        lr.status,
        lr.admin_approval,
        lr.employee_verification,
        lr.created_at
    FROM leave_requests lr
    JOIN users u ON lr.user_id = u.id
    WHERE 1=1
    """
    
    params = []
    
    if status and status != "All":
        query += " AND lr.status = ?"
        params.append(status)
    
    if start_date and end_date:
        query += " AND lr.created_at BETWEEN ? AND ?"
        params.extend([start_date, end_date])
    
    query += " ORDER BY lr.created_at DESC"
    
    df = pd.read_sql_query(query, conn, params=params)
    conn.close()
    return df

def get_project_list():
    conn = sqlite3.connect("attendance_system.db")
    c = conn.cursor()
    c.execute("SELECT name FROM projects")
    projects = [row[0] for row in c.fetchall()]
    conn.close()
    projects.append("General")
    return projects
def get_current_ip():
    """Get current machine IP"""
    s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
    try:
        s.connect(('10.255.255.255', 1))
        ip = s.getsockname()[0]
    except Exception:
        ip = '127.0.0.1'
    finally:
        s.close()
    return ip
def show_announcement_popup():
    if 'dismissed_announcements' not in st.session_state:
        st.session_state.dismissed_announcements = set()

    if 'announcement_timer' not in st.session_state:
        st.session_state.announcement_timer = {}

    conn = sqlite3.connect('attendance_system.db')
    c = conn.cursor()
    c.execute("SELECT id, title, message, created_at FROM announcements ORDER BY created_at DESC LIMIT 1")
    announcement = c.fetchone()
    conn.close()

    if announcement:
        announcement_id, title, message, created_at = announcement

        if announcement_id in st.session_state.dismissed_announcements:
            return

        if announcement_id not in st.session_state.announcement_timer:
            st.session_state.announcement_timer[announcement_id] = time.time()

        if time.time() - st.session_state.announcement_timer[announcement_id] > 60:
            st.session_state.dismissed_announcements.add(announcement_id)
            return

        html_content = f"""
        <div id="popup" style="
            position: fixed;
            bottom: 30px;
            right: 30px;
            background-color: #a245b5;
            color: white;
            padding: 15px;
            border-radius: 10px;
            box-shadow: 0 4px 10px rgba(0,0,0,0.2);
            z-index: 9999;
            width: 300px;
            font-family: Arial, sans-serif;
            animation: fadeInUp 0.5s ease;
        ">
            <div style="display: flex; justify-content: space-between;">
                <strong>📢 New Announcement</strong>
                <button onclick="document.getElementById('popup').remove();" 
                        style="background: transparent; color: white; border: none; font-size: 20px; cursor: pointer;">×</button>
            </div>
            <div style="margin-top: 10px;">
                <b>{title}</b><br>
                🕒 <i>{created_at}</i><br>
                {message}
            </div>
        </div>

        <style>
        @keyframes fadeInUp {{
            from {{ opacity: 0; transform: translateY(20px); }}
            to {{ opacity: 1; transform: translateY(0); }}
        }}
        </style>
        """

        components.html(html_content, height=200)
def generate_qr_code():
    """Generate dynamic QR code with time-based token"""
    current_ip = get_current_ip()
    totp = pyotp.TOTP(pyotp.random_base32(), interval=300)  
    qr = QRCode(f"http://{current_ip}:8501/?token={totp.now()}")
    return qr.png_as_base64_str(scale=6), totp.secret
def get_todays_attendance():
    """Get all attendance records for today"""
    conn = sqlite3.connect('attendance_system.db')
    today = datetime.now().strftime("%Y-%m-%d")
    
    query = """
    SELECT u.name, a.check_in, a.check_out, a.work_hours
    FROM attendance a
    JOIN users u ON a.user_id = u.id
    WHERE a.date = ?
    ORDER BY u.name
    """
    
    df = pd.read_sql_query(query, conn, params=(today,))
    conn.close()
    
    # Format work hours if they exist
    if 'work_hours' in df.columns:
        df['work_hours'] = df['work_hours'].apply(
            lambda x: f"{int(x)}:{int((x % 1) * 60):02d}" if pd.notnull(x) else "N/A"
        )
    
    return df
def qr_attendance_page():
    st.title("Mobile Attendance")

    token = st.query_params.get("token", "")

    totp = pyotp.TOTP(st.secrets["qr_secret"]) 
    if not totp.verify(token, valid_window=1):
        st.error("Invalid or expired QR code")
        return

    auth_method = st.radio("Choose login method:", ["Face Recognition", "Password"])
    
    if auth_method == "Face Recognition":
        img_file = st.camera_input("Face verification")
        if img_file:
            image = Image.open(img_file)
            image_array = np.array(image)
            
            face_encodings_db = load_face_encodings()
            current_encoding = face_recognition.face_encodings(image_array)
            
            if current_encoding:
                matches = face_recognition.compare_faces(
                    list(face_encodings_db.values()), 
                    current_encoding[0]
                )
                if any(matches):
                    user_id = list(face_encodings_db.keys())[matches.index(True)]
                    st.success(f"Welcome User {user_id}!")
                    message = mark_attendance(user_id, "QR Code")
                    st.success(message)
    
    elif auth_method == "Password":
        username = st.text_input("Username")
        password = st.text_input("Password", type="password")
        if st.button("Login"):
            user = authenticate(username, password)
            if user:
                message = mark_attendance(user['id'], "QR Code")
                st.success(f"Welcome {user['name']}! {message}")

if not st.session_state.authenticated and "attendance" in st.query_params:
    qr_attendance_page()

if __name__ == "__main__":
    main()